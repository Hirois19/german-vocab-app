"""Audit the master Excel file for format consistency and data quality.

Checks:
  1. Per-level fill rate (JA / EN / Tags)
  2. Noun completeness (Article + Plural)
  3. Verb completeness (Präteritum + Partizip II)
  4. Garbage in German column (e.g. "1A", "12b", pure numbers, identifiers)

Excludes legitimately-empty cases (uncountable nouns, months, plural-only nouns).
"""

from __future__ import annotations
import sys
import io
import re
from pathlib import Path
from collections import defaultdict

from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

MASTER = Path(__file__).resolve().parent / "german_vocab_master.xlsx"
wb = load_workbook(MASTER)
ws = wb["Session Vokabeln"]

# Nouns that legitimately have no plural (uncountable / abstract / months)
EMPTY_PLURAL_OK_BARE = {
    "Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August",
    "September", "Oktober", "November", "Dezember",
    "Fleisch", "Milch", "Reis", "Zucker", "Butter", "Obst", "Wetter", "Regen",
    "Schnee", "Gesundheit", "Kleidung", "Unterricht", "Fieber", "Husten",
    "Salz", "Pfeffer", "Honig", "Knoblauch", "Eis", "Mehl", "Wasser", "Gold",
    "Silber", "Geld", "Glück", "Pech", "Liebe", "Hass", "Hilfe", "Sport",
    "Musik", "Information", "Verkehr", "Müll", "Strom", "Gas", "Luft", "Sonne",
    "Mond", "Frieden", "Stress", "Trauer", "Wut", "Angst", "Scham", "Stolz",
    "Hoffnung", "Mitleid", "Empathie", "Selbstbewusstsein", "Bewusstsein",
    "Bildung", "Wissen", "Mathematik", "Physik", "Chemie", "Biologie",
    "Medizin", "Psychologie", "Soziologie", "Philosophie", "Forschung",
    "Politik", "Wirtschaft", "Heimat", "Demokratie", "Justiz", "Polizei",
    "Reichtum", "Wohlstand", "Armut", "Schutz", "Recycling", "Verbrauch",
    "Konsum", "Globalisierung", "Digitalisierung", "Software", "Hardware",
    "Automatisierung", "Tatkraft", "Gerechtigkeit", "Solidarität", "Toleranz",
    "Pluralismus", "Nachhaltigkeit", "Verzeihung", "Pflege", "Werbung",
    "Reklame", "Karriere", "Erfahrung", "Wirklichkeit", "Realität", "Wahrheit",
    "Vergangenheit", "Zukunft", "Gegenwart", "Mitte", "Hitze", "Kälte",
    "Sehnsucht", "Einsamkeit", "Korruption", "Bestechung", "Marktforschung",
    "Innenpolitik", "Außenpolitik", "Sozialpolitik", "Wirtschaftspolitik",
    "Bildungspolitik", "Umweltpolitik", "Geburtenrate", "Sterblichkeitsrate",
    "Alterung", "Überalterung", "Asyl", "Pressefreiheit", "Meinungsfreiheit",
    "Gleichberechtigung", "Gleichstellung", "Geschlechtergerechtigkeit",
    "Etablierung", "Konsolidierung", "Stagnation", "Stabilisierung",
    "Destabilisierung", "Deeskalation", "Beilegung", "Schlichtung",
    "Vermittlung", "Komplexität", "Schlichtheit", "Geringfügigkeit",
    "Beträchtlichkeit", "Souveränität", "Eigenständigkeit", "Reichweite",
    "Tragweite", "Wahrnehmungsfähigkeit", "Hingabe", "Unterbewusstsein",
    "Präzision", "Mülltrennung", "Plastikmüll", "Erderwärmung", "Klimawandel",
    "Atomkraft", "Kernenergie", "Wasserkraft", "Windkraft", "Lernen",
    "Statistik", "Inflation", "Wachstum", "Umsatz", "Wettbewerb", "Konkurrenz",
    "Recherche", "Folge", "Lüge", "Diebstahl", "Betrug", "Mord", "Eifersucht",
    "Spielraum", "Aspekt", "Hinsicht", "Gesichtspunkt", "Höhe", "Tiefe",
    "Breite",
}

EMPTY_PLURAL_OK_PHRASES = {
    "die Eltern", "die Geschwister", "die Großeltern", "die Ferien",
    "die Pommes", "die Nudeln", "die Nebenkosten", "die Möbel", "die Medien",
    "die Treibhausgase", "die Daten", "die Überstunden", "die Nachrichten",
}


def looks_garbage(term: str) -> bool:
    """Return True if the German column value is not a real German term."""
    t = term.strip()
    if not t:
        return True
    # Pattern: '1A', '12b', etc.
    if re.fullmatch(r"\d+[a-zA-Z]?", t):
        return True
    # Pattern: 'A1', 'B2' (CEFR-like) standalone
    if re.fullmatch(r"[A-Z]\d", t):
        return True
    # ALL CAPS short (likely page headers)
    if len(t) <= 4 and t.isupper() and t.isalpha():
        return True
    # Contains only punctuation/symbols
    if re.fullmatch(r"[\W_]+", t):
        return True
    # Looks like a section marker
    if re.fullmatch(r"(Kapitel|Sequenz|Lektion|Übung)\s*\d+", t, re.IGNORECASE):
        return True
    return False


stats = defaultdict(lambda: {
    "total": 0, "has_ja": 0, "has_en": 0, "has_tags": 0,
    "noun": 0, "noun_full": 0, "verb": 0, "verb_full": 0,
    "garbage": [], "noun_issues": [], "verb_issues": [],
})

for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if not row or not row[0]:
        continue
    german = str(row[0]).strip()
    pos = str(row[1]).strip().lower() if row[1] else ""
    article = row[2]
    ja = row[3]
    en = row[4]
    prat = row[6] if len(row) > 6 else None
    pii = row[7] if len(row) > 7 else None
    plural = row[8] if len(row) > 8 else None
    tags = row[10] if len(row) > 10 else None

    level = "unknown"
    if tags:
        for t in str(tags).split(","):
            t = t.strip()
            if t in ("A1", "A2", "B1", "B2", "C1"):
                level = t
                break

    s = stats[level]
    s["total"] += 1
    if ja:
        s["has_ja"] += 1
    if en:
        s["has_en"] += 1
    if tags:
        s["has_tags"] += 1

    if looks_garbage(german):
        s["garbage"].append((ridx, german, ja, en))

    if pos == "noun":
        s["noun"] += 1
        bare = re.sub(r"^(der|die|das)\s+", "", german).strip()
        plural_ok = bool(plural) or bare in EMPTY_PLURAL_OK_BARE or german in EMPTY_PLURAL_OK_PHRASES
        article_ok = bool(article)
        if plural_ok and article_ok:
            s["noun_full"] += 1
        else:
            s["noun_issues"].append((ridx, german, article, plural))
    elif pos == "verb":
        s["verb"] += 1
        if prat and pii:
            s["verb_full"] += 1
        else:
            s["verb_issues"].append((ridx, german, prat, pii))

# Report
print("=" * 90)
print("Format consistency per level (excluding legitimately-empty cases)")
print("=" * 90)
hdr = f"{'level':>6} {'total':>6} {'JA%':>5} {'EN%':>5} {'Tags%':>6}    nouns        verbs       garbage"
print(hdr)
print("-" * 90)
for lvl in ["A1", "A2", "B1", "B2", "C1"]:
    s = stats[lvl]
    if s["total"] == 0:
        continue
    ja_pct = s["has_ja"] / s["total"] * 100
    en_pct = s["has_en"] / s["total"] * 100
    tag_pct = s["has_tags"] / s["total"] * 100
    nfp = (s["noun_full"] / s["noun"] * 100) if s["noun"] else 100
    vfp = (s["verb_full"] / s["verb"] * 100) if s["verb"] else 100
    print(
        f"{lvl:>6} {s['total']:>6} {ja_pct:>4.0f}% {en_pct:>4.0f}% {tag_pct:>5.0f}%   "
        f"{s['noun_full']:>3}/{s['noun']:<3}({nfp:>3.0f}%)  "
        f"{s['verb_full']:>3}/{s['verb']:<3}({vfp:>3.0f}%)  "
        f"{len(s['garbage']):>3}"
    )

print()
print("=" * 90)
print("Garbage entries in German column (e.g. '1A', '12b', identifiers, pure numbers)")
print("=" * 90)
total_g = 0
for lvl in ["A1", "A2", "B1", "B2", "C1"]:
    g = stats[lvl]["garbage"]
    if not g:
        continue
    print(f"\n[{lvl}] {len(g)} garbage entries:")
    for ridx, german, ja, en in g:
        print(f"  row {ridx}: German=\"{german}\" | JA={ja} | EN={en}")
    total_g += len(g)
if total_g == 0:
    print("\n  None found.")
else:
    print(f"\nTotal garbage entries: {total_g}")

print()
print("=" * 90)
print("Noun issues (missing Article or Plural, excluding uncountables)")
print("=" * 90)
total_n = 0
for lvl in ["A1", "A2", "B1", "B2", "C1"]:
    issues = stats[lvl]["noun_issues"]
    if not issues:
        continue
    print(f"\n[{lvl}] {len(issues)} noun issues:")
    for ridx, german, article, plural in issues[:30]:
        print(f"  row {ridx}: \"{german}\" article={article!r}, plural={plural!r}")
    if len(issues) > 30:
        print(f"  ... and {len(issues)-30} more")
    total_n += len(issues)
if total_n == 0:
    print("\n  None.")

print()
print("=" * 90)
print("Verb issues (missing Präteritum or Partizip II)")
print("=" * 90)
total_v = 0
for lvl in ["A1", "A2", "B1", "B2", "C1"]:
    issues = stats[lvl]["verb_issues"]
    if not issues:
        continue
    print(f"\n[{lvl}] {len(issues)} verb issues:")
    for ridx, german, prat, pii in issues[:30]:
        print(f"  row {ridx}: \"{german}\" prät={prat!r}, pii={pii!r}")
    if len(issues) > 30:
        print(f"  ... and {len(issues)-30} more")
    total_v += len(issues)
if total_v == 0:
    print("\n  None.")
