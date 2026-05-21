"""One-shot script: add a `Tags` column to german_vocab_session_new.xlsx.

For each row in the master, compute and write tags directly into the Excel:
  - Theme tags (Familie, Essen, Arbeit, …) via keyword matching against
    German headword + EN/JA translations.
  - POS tag (verb / noun / adj / adv / conj / …) from the existing POS column.
  - CEFR level tag — session_new vocab is all B1.

After running this script, the Excel is self-contained: opening it shows the
tags for each word in column K (and they survive future edits unless the user
manually changes them).

The extract pipeline (extract_vocab.py) will read this column directly instead
of recomputing tags at extract time.
"""

from __future__ import annotations
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

SRC = Path(__file__).resolve().parent / "german_vocab_master.xlsx"
SHEET = "Session Vokabeln"
LEVEL = "B1"  # session_new is sourced from a B1 Intensivkurs


# Theme keywords — mirrored from extract_vocab.py THEMES.
THEMES: dict[str, list[str]] = {
    "Familie": [
        "mother", "father", "brother", "sister", "wife", "husband", "family",
        "parent", "child", "son", "daughter", "uncle", "aunt", "grandmother",
        "grandfather", "cousin", "nephew", "niece",
        "母", "父", "兄", "姉", "弟", "妹", "夫", "妻", "家族",
        "両親", "祖", "息子", "娘", "親戚",
        "Mutter", "Vater", "Bruder", "Schwester", "Eltern", "Familie",
        "Kind", "Sohn", "Tochter", "Onkel", "Tante", "Oma", "Opa",
        "Großmutter", "Großvater", "Großeltern", "Geschwister", "Enkel",
        "Baby", "Cousin", "Cousine", "Verwandte", "Verwandter", "Ehepaar",
        "Ehefrau", "Ehemann", "Schwager", "Schwägerin", "Schwiegermutter",
    ],
    "Essen": [
        "food", "eat", "drink", "bread", "meat", "fish", "vegetable", "fruit",
        "cheese", "milk", "coffee", "tea", "wine", "beer", "water", "soup",
        "breakfast", "lunch", "dinner", "restaurant", "cook", "kitchen",
        "食べ", "飲み", "パン", "肉", "魚", "野菜", "果物", "コーヒー", "茶", "水",
        "朝食", "昼食", "夕食", "レストラン", "料理", "台所", "牛乳", "ワイン", "ビール",
        "essen", "Essen", "trinken", "Brot", "Brötchen", "Butter", "Käse",
        "Wurst", "Schinken", "Fleisch", "Fisch", "Gemüse", "Obst", "Apfel",
        "Banane", "Orange", "Tomate", "Kartoffel", "Zwiebel", "Salat",
        "Suppe", "Reis", "Nudel", "Pasta", "Kuchen", "Schokolade", "Bonbon",
        "Salz", "Pfeffer", "Zucker", "Senf", "Saft", "Wasser", "Milch",
        "Kaffee", "Tee", "Wein", "Bier", "Glas", "Tasse", "Teller", "Gabel",
        "Messer", "Löffel", "kochen", "backen", "braten", "Restaurant",
        "Café", "Bäckerei", "Metzgerei", "Frühstück", "Mittagessen",
        "Abendessen", "Mahlzeit", "Hunger", "Durst", "satt", "lecker",
        "schmecken", "Geschmack", "Imbiss",
    ],
    "Arbeit": [
        "work", "job", "office", "boss", "colleague", "employee", "salary",
        "meeting", "company", "career", "profession", "business",
        "仕事", "会社", "オフィス", "同僚", "給料", "会議", "ビジネス", "職業", "上司",
        "Arbeit", "arbeiten", "Beruf", "beruflich", "Job", "Büro", "Firma",
        "Unternehmen", "Chef", "Chefin", "Kollege", "Kollegin", "Mitarbeiter",
        "Angestellter", "Angestellte", "Lohn", "Gehalt", "Karriere",
        "Konferenz", "Besprechung", "Sitzung", "Projekt", "Kunde", "Kundin",
        "Geschäft", "Stelle", "Bewerbung", "Bewerber", "Bewerberin",
        "Vorstellungsgespräch", "Aufgabe", "Termin", "kündigen", "anstellen",
        "einstellen", "Beförderung", "Ausbildung", "Praktikum",
    ],
    "Reisen": [
        "travel", "trip", "holiday", "vacation", "airport", "flight", "hotel",
        "passport", "ticket", "luggage", "tourist",
        "旅", "旅行", "休暇", "空港", "飛行", "ホテル", "パスポート", "観光",
        "Reise", "reisen", "Urlaub", "Ferien", "Flug", "fliegen", "Flughafen",
        "Hotel", "Unterkunft", "Pension", "Ticket", "Pass", "Reisepass",
        "Koffer", "Gepäck", "Tourist", "Touristin", "Sehenswürdigkeit",
        "Ausflug", "Sightseeing", "Visum", "Grenze", "Zoll", "Souvenir",
    ],
    "Verkehr": [
        "car", "bus", "train", "bike", "taxi", "drive", "street", "road",
        "traffic", "station", "platform",
        "車", "バス", "電車", "自転車", "タクシー", "運転", "駅",
        "Auto", "Wagen", "Bus", "Zug", "Bahn", "Straßenbahn", "U-Bahn",
        "S-Bahn", "Fahrrad", "Rad", "Motorrad", "Taxi", "fahren", "Fahrer",
        "Reifen", "Tankstelle", "Benzin", "Diesel", "Straße", "Weg",
        "Autobahn", "Bahnhof", "Haltestelle", "Gleis", "Ampel", "Kreuzung",
        "Verkehr", "Stau", "parken", "Parkplatz", "Führerschein",
    ],
    "Wohnen": [
        "house", "apartment", "flat", "room", "kitchen", "bathroom", "bedroom",
        "garden", "rent", "furniture", "door", "window", "wall",
        "家", "部屋", "アパート", "キッチン", "風呂", "庭", "家賃", "家具", "ドア", "窓",
        "Haus", "Wohnung", "Zimmer", "Küche", "Bad", "Schlafzimmer",
        "Wohnzimmer", "Esszimmer", "Garten", "Balkon", "Terrasse", "Möbel",
        "Tisch", "Stuhl", "Sofa", "Bett", "Schrank", "Regal", "Lampe", "Tür",
        "Fenster", "Wand", "Boden", "Decke", "Dach", "Miete", "wohnen",
        "Mieter", "Vermieter", "Nachbar", "Adresse",
    ],
    "Gesundheit": [
        "health", "sick", "doctor", "hospital", "medicine", "pain",
        "fever", "illness", "pharmacy",
        "健康", "病気", "医者", "病院", "痛",
        "Gesundheit", "gesund", "krank", "Krankheit", "Arzt", "Ärztin",
        "Krankenhaus", "Klinik", "Apotheke", "Apotheker", "Medikament",
        "Tablette", "Pille", "Schmerz", "Fieber", "Husten", "Erkältung",
        "Grippe", "Verletzung", "Wunde", "Operation", "Therapie", "Patient",
        "Symptom", "Diagnose", "Rezept", "Heilung", "Blut", "Puls",
    ],
    "Kleidung": [
        "clothes", "shirt", "trouser", "trousers", "dress", "shoe", "shoes",
        "jacket", "coat", "skirt", "hat", "sock", "socks",
        "シャツ", "ズボン", "靴", "ジャケット", "コート", "スカート", "帽子",
        "Kleidung", "Kleid", "Hemd", "Bluse", "Pullover", "Pulli", "Jacke",
        "Mantel", "Hose", "Rock", "Anzug", "Schuh", "Stiefel", "Sandale",
        "Socke", "Strumpf", "Mütze", "Hut", "Schal", "Handschuh", "Gürtel",
        "Tasche", "anziehen", "ausziehen", "Krawatte",
    ],
    "Einkaufen": [
        "shop", "store", "buy", "sell", "pay", "money", "price", "cheap",
        "expensive", "supermarket",
        "店", "買う", "売る", "支払", "値段", "安い", "高い", "スーパー",
        "einkaufen", "kaufen", "Kauf", "Supermarkt", "Geschäft", "Laden",
        "Markt", "Einkauf", "Preis", "kosten", "billig", "teuer", "Euro",
        "Geld", "Bargeld", "Quittung", "Rechnung", "bezahlen", "zahlen",
        "ausgeben", "sparen", "Kasse", "Verkäufer", "Rabatt", "Angebot",
    ],
    "Zeit": [
        "hour", "minute", "day", "week", "month", "year", "morning",
        "evening", "night", "today", "tomorrow", "yesterday",
        "時間", "分", "週", "月", "年", "朝", "夕方", "夜", "今日", "明日", "昨日",
        "Zeit", "Stunde", "Minute", "Sekunde", "Tag", "Woche", "Monat",
        "Jahr", "Wochenende", "heute", "morgen", "gestern", "vorgestern",
        "übermorgen", "jetzt", "später", "früher", "früh", "spät",
        "Termin", "Datum", "Kalender", "Montag", "Dienstag", "Mittwoch",
        "Donnerstag", "Freitag", "Samstag", "Sonntag", "Januar", "Februar",
        "März", "April", "Juni", "Juli", "August", "September", "Oktober",
        "November", "Dezember", "Morgen", "Mittag", "Abend", "Nacht",
    ],
    "Wetter": [
        "weather", "rain", "snow", "sun", "wind", "cloud", "cold", "hot",
        "warm", "temperature", "storm",
        "天気", "雨", "雪", "太陽", "寒い", "暑い", "暖かい", "気温",
        "Wetter", "Regen", "regnen", "Schnee", "schneien", "Sonne", "sonnig",
        "Wind", "windig", "Wolke", "wolkig", "Sturm", "Gewitter", "Klima",
        "kalt", "warm", "heiß", "kühl", "frieren", "schwitzen", "Temperatur",
        "Grad", "Hagel", "Nebel", "neblig",
    ],
    "Schule": [
        "school", "teacher", "student", "class", "lesson", "learn", "study",
        "homework", "exam", "university",
        "学校", "先生", "学生", "授業", "勉強", "宿題", "試験", "大学",
        "Schule", "Schüler", "Schülerin", "Lehrer", "Lehrerin", "Lehrkraft",
        "Klasse", "Unterricht", "lernen", "studieren", "Studium", "Student",
        "Studentin", "Hausaufgabe", "Prüfung", "Klausur", "Test", "Note",
        "Zeugnis", "Universität", "Hochschule", "Fach", "Mathematik",
        "Deutsch", "Englisch", "Geschichte", "Biologie", "Chemie", "Physik",
    ],
    "Freizeit": [
        "sport", "sports", "game", "play", "music", "film", "movie", "book",
        "read", "hobby", "party", "dance",
        "スポーツ", "ゲーム", "音楽", "映画", "本", "趣味", "パーティー", "ダンス",
        "Sport", "Spiel", "spielen", "Freizeit", "Hobby", "Musik", "Film",
        "Kino", "Theater", "Konzert", "Museum", "feiern", "Fest", "tanzen",
        "Disko", "Tanz", "Schwimmen", "Joggen", "Wandern", "Camping",
        "Picknick", "Spaziergang", "Spaß", "Vergnügen",
    ],
    "Gefühle": [
        "happy", "sad", "angry", "afraid", "love", "feel", "feeling",
        "嬉", "悲", "怒", "怖", "愛", "感じ",
        "Gefühl", "fühlen", "glücklich", "Glück", "traurig", "Trauer",
        "wütend", "Wut", "ängstlich", "Angst", "lieben", "Liebe", "hassen",
        "Hass", "mögen", "Lust", "Freude", "freuen", "Spaß", "hoffen",
        "Hoffnung", "Sorge", "Stress", "gestresst", "ruhig", "nervös",
    ],
    "Farben": [
        "color", "colour", "red", "blue", "green", "yellow", "black", "white",
        "brown", "pink", "orange", "grey", "gray",
        "色", "赤", "青", "緑", "黄色", "黒", "白",
        "Farbe", "rot", "blau", "grün", "gelb", "schwarz", "weiß", "braun",
        "grau", "lila", "rosa", "türkis", "beige", "golden", "silbern",
        "dunkel", "hell", "bunt",
    ],
}

_ASCII_WORD = re.compile(r"^[a-zA-Z]+$")


def has_keyword(haystack: str, keyword: str) -> bool:
    if _ASCII_WORD.fullmatch(keyword):
        return bool(re.search(rf"\b{re.escape(keyword)}\b", haystack, re.IGNORECASE))
    return keyword in haystack


def compute_tags(german: str, pos: str | None, ja: str | None, en: str | None) -> list[str]:
    haystack = " ".join(filter(None, [german, ja or "", en or ""]))
    tags: list[str] = []
    for theme, kws in THEMES.items():
        for kw in kws:
            if has_keyword(haystack, kw):
                tags.append(theme)
                break
    if pos:
        tags.append(pos.strip().lower())
    tags.append(LEVEL)
    # Dedup, preserve order
    seen: set[str] = set()
    out = []
    for t in tags:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


def main():
    wb = load_workbook(SRC)
    ws = wb[SHEET]
    header = [c.value for c in ws[1]]
    print(f"Existing header: {header}")

    # Determine where to write Tags. If "Tags" already exists, overwrite.
    tags_col_idx = None
    for i, h in enumerate(header):
        if h == "Tags":
            tags_col_idx = i + 1
            break
    if tags_col_idx is None:
        tags_col_idx = len(header) + 1
        ws.cell(row=1, column=tags_col_idx, value="Tags")
        print(f"Added Tags column at position {tags_col_idx}")
    else:
        print(f"Overwriting existing Tags column at position {tags_col_idx}")

    rows_written = 0
    for row in ws.iter_rows(min_row=2):
        # Read German, POS, Japanese, English (cols A, B, D, E)
        german = row[0].value
        pos = row[1].value
        ja = row[3].value
        en = row[4].value
        if not german or not str(german).strip():
            continue
        tags = compute_tags(str(german), str(pos) if pos else None,
                            str(ja) if ja else None, str(en) if en else None)
        row[tags_col_idx - 1].value = ", ".join(tags)
        rows_written += 1

    wb.save(SRC)
    print(f"Wrote tags for {rows_written} rows.")
    print(f"Saved to {SRC}")


if __name__ == "__main__":
    main()
