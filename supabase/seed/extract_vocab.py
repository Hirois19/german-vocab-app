"""
Extract unique German vocabulary from source files into a single vocab.json.

Sources (12 files in 05_reference/misc/German/):
- B1_0001-0200.xlsx 〜 B1_1201-1400.xlsx : 7 files, B1 graded (Der/Die/Das | term | JA, paired 2-column layout)
- German Vocablary.xlsx : 7 sheets, 200 words each (番号 | Der/Die/Das | term | JA | SEKI 1-7 cols)
- german_vocab_session_new.xlsx : 1 sheet, rich schema (POS | Article | JA | EN | Präteritum | Partizip II | Plural | Notes)
- NWn_A1_Glossar_Deutsch-Englisch.pdf : A1 glossary (line-based, EN translation)
- NWn_A2_Glossar_Englisch.pdf : skipped (layout corrupted)
- NWn_B1_Glossar_Englisch.pdf : skipped (redundant with Excel, layout corrupted)

Output:
- data/vocab.json : deduplicated unique vocabulary with merged metadata
- data/vocab.csv  : same data in CSV for human review

Dedup key: normalized form of the German term
  - articles removed (der/die/das/den/dem/des/ein/eine/etc.)
  - lowercased
  - whitespace collapsed
  - separable verb pipe removed (zu|hören -> zuhören)
"""
from __future__ import annotations

import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
import pymupdf as fitz

sys.stdout.reconfigure(encoding="utf-8")

# Paths — everything is resolved relative to this script so the pipeline is
# reproducible from a fresh clone.
SEED_DIR = Path(__file__).resolve().parent
SRC_DIR = SEED_DIR
OUT_DIR = SEED_DIR
OUT_JSON = OUT_DIR / "vocab.json"
OUT_CSV = OUT_DIR / "vocab.csv"

ARTICLE_RE = re.compile(
    r"^(der|die|das|den|dem|des|ein|eine|einen|einem|einer|eines)\s+",
    re.IGNORECASE,
)
ARTICLE_VALUES = {"der", "die", "das"}


def normalize_key(term: str) -> str:
    if not term:
        return ""
    s = str(term).strip()
    s = re.sub(r"\s+", " ", s)
    s = ARTICLE_RE.sub("", s)
    s = s.replace("|", "")
    s = s.lower()
    return s.strip()


# Theme keyword groups. When a card matches a theme (via translations_en or
# translations_ja), the theme name is added to `entry["categories"]` so the
# shared cards table can be queried by theme without runtime matching. The app
# reads these via `listAvailableTags()` in lib/db/cards.ts.
THEMES: dict[str, list[str]] = {
    "Familie": [
        # English
        "mother", "father", "brother", "sister", "wife", "husband", "family",
        "parent", "child", "son", "daughter", "uncle", "aunt", "grandmother",
        "grandfather", "cousin", "nephew", "niece",
        # Japanese
        "母", "父", "兄", "姉", "弟", "妹", "夫", "妻", "家族",
        "両親", "祖", "息子", "娘", "親戚",
        # German (head words)
        "Mutter", "Vater", "Bruder", "Schwester", "Eltern", "Familie",
        "Kind", "Sohn", "Tochter", "Onkel", "Tante", "Oma", "Opa",
        "Großmutter", "Großvater", "Großeltern", "Geschwister", "Enkel",
        "Baby", "Cousin", "Cousine", "Verwandte", "Verwandter", "Ehepaar",
        "Ehefrau", "Ehemann", "Schwager", "Schwägerin", "Schwiegermutter",
        "Schwiegervater", "Schwiegersohn", "Schwiegertochter",
    ],
    "Essen": [
        # English
        "food", "eat", "drink", "bread", "meat", "fish", "vegetable", "fruit",
        "cheese", "milk", "coffee", "tea", "wine", "beer", "water", "soup",
        "breakfast", "lunch", "dinner", "restaurant", "cook", "kitchen",
        # Japanese
        "食べ", "飲み", "パン", "肉", "魚", "野菜", "果物", "コーヒー", "茶", "水",
        "朝食", "昼食", "夕食", "レストラン", "料理", "台所", "牛乳", "ワイン", "ビール",
        # German
        "essen", "Essen", "trinken", "Brot", "Brötchen", "Butter", "Käse",
        "Wurst", "Schinken", "Fleisch", "Fisch", "Gemüse", "Obst", "Apfel",
        "Banane", "Orange", "Tomate", "Kartoffel", "Zwiebel", "Salat",
        "Suppe", "Reis", "Nudel", "Pasta", "Kuchen", "Schokolade", "Bonbon",
        "Salz", "Pfeffer", "Zucker", "Senf", "Saft", "Wasser", "Milch",
        "Kaffee", "Tee", "Wein", "Bier", "Glas", "Tasse", "Teller", "Gabel",
        "Messer", "Löffel", "kochen", "backen", "braten", "Restaurant",
        "Café", "Bäckerei", "Metzgerei", "Frühstück", "Mittagessen",
        "Abendessen", "Mahlzeit", "Hunger", "Durst", "satt", "lecker",
        "schmecken", "Geschmack", "Imbiss",
    ],
    "Arbeit": [
        "work", "job", "office", "boss", "colleague", "employee", "salary",
        "meeting", "company", "career", "profession", "business",
        "仕事", "会社", "オフィス", "同僚", "給料", "会議", "ビジネス", "職業", "上司",
        "Arbeit", "arbeiten", "Beruf", "beruflich", "Job", "Büro", "Firma",
        "Unternehmen", "Chef", "Chefin", "Kollege", "Kollegin", "Mitarbeiter",
        "Angestellter", "Angestellte", "Lohn", "Gehalt", "Karriere",
        "Konferenz", "Besprechung", "Sitzung", "Projekt", "Kunde", "Kundin",
        "Geschäft", "Stelle", "Bewerbung", "Bewerber", "Bewerberin",
        "Vorstellungsgespräch", "Aufgabe", "Termin", "kündigen", "anstellen",
        "einstellen", "Beförderung", "Gehalt", "Urlaub", "Pause",
        "Berufsausbildung", "Ausbildung", "Praktikum", "Werkstatt",
    ],
    "Reisen": [
        "travel", "trip", "holiday", "vacation", "airport", "flight", "hotel",
        "passport", "ticket", "luggage", "tourist",
        "旅", "旅行", "休暇", "空港", "飛行", "ホテル", "パスポート", "観光",
        "Reise", "reisen", "Urlaub", "Ferien", "Flug", "fliegen", "Flughafen",
        "Hotel", "Unterkunft", "Pension", "Ticket", "Pass", "Reisepass",
        "Koffer", "Gepäck", "Tourist", "Touristin", "Sehenswürdigkeit",
        "Ausflug", "Sightseeing", "Visum", "Grenze", "Zoll", "Souvenir",
        "Andenken", "Reiseführer", "Reisebüro", "Strand", "Berg", "Berge",
    ],
    "Verkehr": [
        "car", "bus", "train", "bike", "taxi", "drive", "street", "road",
        "traffic", "station", "platform",
        "車", "バス", "電車", "自転車", "タクシー", "運転", "駅",
        "Auto", "Wagen", "Bus", "Zug", "Bahn", "Straßenbahn", "U-Bahn",
        "S-Bahn", "Fahrrad", "Rad", "Motorrad", "Taxi", "fahren", "Fahrer",
        "Fahrerin", "Reifen", "Tankstelle", "Benzin", "Diesel", "Straße",
        "Weg", "Autobahn", "Bahnhof", "Haltestelle", "Gleis", "Ampel",
        "Kreuzung", "Verkehr", "Stau", "parken", "Parkplatz", "Führerschein",
    ],
    "Wohnen": [
        "house", "apartment", "flat", "room", "kitchen", "bathroom", "bedroom",
        "garden", "rent", "furniture", "door", "window", "wall",
        "家", "部屋", "アパート", "キッチン", "風呂", "庭", "家賃", "家具", "ドア", "窓",
        "Haus", "Wohnung", "Zimmer", "Küche", "Bad", "Schlafzimmer",
        "Wohnzimmer", "Esszimmer", "Garten", "Balkon", "Terrasse", "Möbel",
        "Tisch", "Stuhl", "Sofa", "Bett", "Schrank", "Regal", "Lampe", "Tür",
        "Fenster", "Wand", "Boden", "Decke", "Dach", "Miete", "wohnen",
        "Mieter", "Vermieter", "Vermieterin", "Nachbar", "Nachbarin",
        "Adresse", "Postleitzahl", "Spüle", "Dusche", "Heizung", "Kühlschrank",
    ],
    "Gesundheit": [
        "health", "sick", "doctor", "hospital", "medicine", "pain",
        "fever", "illness", "pharmacy",
        "健康", "病気", "医者", "病院", "痛",
        "Gesundheit", "gesund", "krank", "Krankheit", "Arzt", "Ärztin",
        "Krankenhaus", "Klinik", "Apotheke", "Apotheker", "Apothekerin",
        "Medikament", "Tablette", "Pille", "Schmerz", "Fieber", "Husten",
        "Erkältung", "Grippe", "Verletzung", "Wunde", "Operation", "Therapie",
        "Patient", "Patientin", "Symptom", "Diagnose", "Rezept", "Heilung",
        "Verband", "Spritze", "Impfung", "Blut", "Puls",
    ],
    "Kleidung": [
        "clothes", "shirt", "trouser", "trousers", "dress", "shoe", "shoes",
        "jacket", "coat", "skirt", "hat", "sock", "socks",
        "シャツ", "ズボン", "靴", "ジャケット", "コート", "スカート", "帽子",
        "Kleidung", "Kleid", "Hemd", "Bluse", "Pullover", "Pulli", "Jacke",
        "Mantel", "Hose", "Rock", "Anzug", "Schuh", "Stiefel", "Sandale",
        "Socke", "Strumpf", "Mütze", "Hut", "Schal", "Handschuh", "Gürtel",
        "Tasche", "anziehen", "ausziehen", "Krawatte", "Unterwäsche",
    ],
    "Einkaufen": [
        "shop", "store", "buy", "sell", "pay", "money", "price", "cheap",
        "expensive", "supermarket",
        "店", "買う", "売る", "支払", "値段", "安い", "高い", "スーパー",
        "einkaufen", "kaufen", "Kauf", "Supermarkt", "Geschäft", "Laden",
        "Markt", "Einkauf", "Preis", "kosten", "billig", "teuer", "Euro",
        "Geld", "Bargeld", "Quittung", "Rechnung", "bezahlen", "zahlen",
        "ausgeben", "sparen", "Sparen", "Kasse", "Verkäufer", "Verkäuferin",
        "Kunde", "Rabatt", "Angebot",
    ],
    "Zeit": [
        "hour", "minute", "day", "week", "month", "year", "morning",
        "evening", "night", "today", "tomorrow", "yesterday",
        "時間", "分", "週", "月", "年", "朝", "夕方", "夜", "今日", "明日", "昨日",
        "Zeit", "Stunde", "Minute", "Sekunde", "Tag", "Woche", "Monat",
        "Jahr", "Wochenende", "heute", "morgen", "gestern", "vorgestern",
        "übermorgen", "jetzt", "später", "früher", "früh", "spät",
        "Termin", "Datum", "Kalender", "Montag", "Dienstag", "Mittwoch",
        "Donnerstag", "Freitag", "Samstag", "Sonntag", "Januar", "Februar",
        "März", "April", "Juni", "Juli", "August", "September", "Oktober",
        "November", "Dezember", "Morgen", "Mittag", "Abend", "Nacht",
    ],
    "Wetter": [
        "weather", "rain", "snow", "sun", "wind", "cloud", "cold", "hot",
        "warm", "temperature", "storm",
        "天気", "雨", "雪", "太陽", "寒い", "暑い", "暖かい", "気温",
        "Wetter", "Regen", "regnen", "Schnee", "schneien", "Sonne", "sonnig",
        "Wind", "windig", "Wolke", "wolkig", "Sturm", "Gewitter", "Klima",
        "kalt", "warm", "heiß", "kühl", "frieren", "schwitzen", "Temperatur",
        "Grad", "Hagel", "Nebel", "neblig",
    ],
    "Schule": [
        "school", "teacher", "student", "class", "lesson", "learn", "study",
        "homework", "exam", "university",
        "学校", "先生", "学生", "授業", "勉強", "宿題", "試験", "大学",
        "Schule", "Schüler", "Schülerin", "Lehrer", "Lehrerin", "Lehrkraft",
        "Klasse", "Unterricht", "lernen", "studieren", "Studium", "Student",
        "Studentin", "Hausaufgabe", "Prüfung", "Klausur", "Test", "Note",
        "Zeugnis", "Universität", "Hochschule", "Fach", "Mathematik",
        "Deutsch", "Englisch", "Geschichte", "Biologie", "Chemie", "Physik",
        "Tafel", "Heft", "Buch", "Stift", "Bleistift", "Kuli", "Kugelschreiber",
    ],
    "Freizeit": [
        "sport", "sports", "game", "play", "music", "film", "movie", "book",
        "read", "hobby", "party", "dance",
        "スポーツ", "ゲーム", "音楽", "映画", "本", "趣味", "パーティー", "ダンス",
        "Sport", "Spiel", "spielen", "Freizeit", "Hobby", "Musik", "Film",
        "Kino", "Theater", "Konzert", "Museum", "feiern", "Fest", "tanzen",
        "Disko", "Tanz", "Schwimmen", "schwimmen", "Joggen", "joggen",
        "Wandern", "wandern", "Camping", "Picknick", "Spaziergang",
        "spazieren", "Spaß", "Vergnügen", "Karte", "Brettspiel", "Fußball",
        "Tennis", "Basketball",
    ],
    "Gefühle": [
        "happy", "sad", "angry", "afraid", "love", "feel", "feeling",
        "嬉", "悲", "怒", "怖", "愛", "感じ",
        "Gefühl", "fühlen", "glücklich", "Glück", "traurig", "Trauer",
        "wütend", "Wut", "ängstlich", "Angst", "lieben", "Liebe", "hassen",
        "Hass", "mögen", "Lust", "Freude", "freuen", "Spaß", "hoffen",
        "Hoffnung", "Sorge", "sorgen", "Stress", "gestresst", "ruhig",
        "nervös", "stolz", "Stolz", "Scham", "schämen", "lachen", "weinen",
    ],
    "Farben": [
        "color", "colour", "red", "blue", "green", "yellow", "black", "white",
        "brown", "pink", "orange", "grey", "gray",
        "色", "赤", "青", "緑", "黄色", "黒", "白",
        "Farbe", "rot", "blau", "grün", "gelb", "schwarz", "weiß", "braun",
        "grau", "lila", "rosa", "türkis", "beige", "golden", "silbern",
        "dunkel", "hell", "bunt",
    ],
}

_ASCII_WORD = re.compile(r"^[a-zA-Z]+$")


def _has_keyword(haystack: str, keyword: str) -> bool:
    """Word-boundary match for ASCII; substring match for CJK."""
    if _ASCII_WORD.fullmatch(keyword):
        return bool(re.search(rf"\b{re.escape(keyword)}\b", haystack, re.IGNORECASE))
    return keyword in haystack


def compute_categories(entry: dict) -> list[str]:
    """Return the list of theme + level tags this entry matches.

    Matches keywords against the German headword AND the EN/JA translations.
    Additionally appends the card's CEFR level(s) so every entry has at least
    one tag.
    """
    de_term = entry.get("term_de", "")
    translations = " ".join(
        (entry.get("translations_en") or []) + (entry.get("translations_ja") or [])
    )
    haystack = f"{de_term} {translations}"
    out: list[str] = []
    for theme, keywords in THEMES.items():
        for kw in keywords:
            if _has_keyword(haystack, kw):
                out.append(theme)
                break
    # Add CEFR level tags so every card carries at least one tag.
    for level in entry.get("levels") or []:
        if level not in out:
            out.append(level)
    return out


def is_noise(term: str) -> bool:
    """Filter out clearly non-vocabulary entries."""
    if not term:
        return True
    s = str(term).strip()
    if len(s) < 2:
        return True
    if len(s) > 60:
        return True
    # Starts/ends with parenthesis (likely example sentence)
    if s.startswith("(") or s.startswith("[") or s.startswith("\""):
        return True
    # Full sentences (with terminal punctuation)
    if re.search(r"[.!?]\s*$", s) and " " in s:
        return True
    # Pure digits or punctuation
    if re.fullmatch(r"[\d\s\W]+", s):
        return True
    # Starts with a digit (price examples like "15 €", exercise markers like "10a")
    if re.match(r"^\s*\d", s):
        return True
    # Currency or other characters typical of example sentences, not headwords
    if re.search(r"[€$¥¢]", s):
        return True
    # Common header/note tokens that leak from spreadsheet metadata
    if s.lower() in {"none", "null", "nan", "deutsch", "englisch", "japanisch"}:
        return True
    # Section / chapter / test headers leaking from B2/C1 vocab PDFs.
    # Require digit after to avoid killing valid vocab like "Aufgabe" / "Test".
    if re.match(r"^(Test|Aufgabe|Lektion|Kapitel|Modul|Teil|Übung|S\.|Seite)\s+\d", s):
        return True
    # Unmatched closing parenthesis without an opener — broken fragment from
    # a wrapped example sentence (e.g., "werden)", "Arbeit.)").
    if ")" in s and "(" not in s:
        return True
    # Ends with ", -" or just "-" — typically a stray plural marker fragment.
    if re.search(r",\s*[\-\"]?\s*$", s):
        return True
    return False


def add_entry(
    bucket: dict,
    raw_term: str,
    *,
    article: str | None = None,
    pos: str | None = None,
    ja: str | None = None,
    en: str | None = None,
    prateritum: str | None = None,
    partizip_ii: str | None = None,
    plural: str | None = None,
    notes: str | None = None,
    example: str | None = None,
    level: str | None = None,
    categories: list[str] | None = None,
    source: str,
):
    if is_noise(raw_term):
        return
    key = normalize_key(raw_term)
    if not key or key in {"none", "null", "nan"}:
        return
    entry = bucket.setdefault(
        key,
        {
            "canonical_key": key,
            "term_de": str(raw_term).strip(),
            "article": None,
            "pos": None,
            "translations_ja": [],
            "translations_en": [],
            "examples": [],
            "prateritum": None,
            "partizip_ii": None,
            "plural": None,
            "notes": [],
            "levels": [],
            "sources": [],
            "categories": [],
        },
    )

    if article and article.lower() in ARTICLE_VALUES:
        entry["article"] = entry["article"] or article.lower()

    if pos and not entry["pos"]:
        entry["pos"] = pos.lower()

    if ja:
        ja_s = str(ja).strip()
        if ja_s and ja_s not in entry["translations_ja"]:
            entry["translations_ja"].append(ja_s)

    if en:
        en_s = str(en).strip()
        if en_s and en_s not in entry["translations_en"]:
            entry["translations_en"].append(en_s)

    if prateritum and not entry["prateritum"]:
        entry["prateritum"] = str(prateritum).strip()
    if partizip_ii and not entry["partizip_ii"]:
        entry["partizip_ii"] = str(partizip_ii).strip()
    if plural and not entry["plural"]:
        entry["plural"] = str(plural).strip()
    if notes:
        n = str(notes).strip()
        if n and n not in entry["notes"]:
            entry["notes"].append(n)
    if example:
        ex = str(example).strip()
        if ex and ex not in entry["examples"]:
            entry["examples"].append(ex)
    if level and level not in entry["levels"]:
        entry["levels"].append(level)
    if source not in entry["sources"]:
        entry["sources"].append(source)
    if categories:
        for c in categories:
            c = c.strip()
            if c and c not in entry["categories"]:
                entry["categories"].append(c)


# ---------- Parsers ----------

def parse_b1_graded(bucket: dict, path: Path):
    """B1_0001-0200.xlsx etc.
    Schema: row1 header (Der/Die/Das | 単語 | 日本語 | (blank) | Der/Die/Das | 単語 | 日本語)
    Data: 2 words per row (left + right blocks).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Left block
        if row[0] or row[1] or row[2]:
            add_entry(
                bucket,
                row[1],
                article=row[0],
                ja=row[2],
                level="B1",
                source=path.name,
            )
        # Right block (skip blank col 4)
        if len(row) >= 7 and (row[4] or row[5] or row[6]):
            add_entry(
                bucket,
                row[5],
                article=row[4],
                ja=row[6],
                level="B1",
                source=path.name,
            )


def parse_german_vocablary(bucket: dict, path: Path):
    """German Vocablary.xlsx with 7 sheets.
    Schema: 番号 | Der/Die/Das | 単語 | 日本語 | 1.0 | 2.0 | ... | 7.0
    """
    wb = load_workbook(path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[2]:
                continue
            add_entry(
                bucket,
                row[2],
                article=row[1],
                ja=row[3],
                level="B1",
                source=f"{path.name}::{sn}",
            )


def parse_session_new(bucket: dict, path: Path):
    """german_vocab_session_new.xlsx — master vocab list.
    Schema: German | POS | Article | Japanese | English | Appearance_Count |
            Präteritum | Partizip II | Plural | Notes | Tags
    Tags column (added by add_tags_to_session_new.py) is comma-separated and
    is the authoritative tag list for each card — we read it directly instead
    of recomputing at extract time.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb["Session Vokabeln"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        tags_str = row[10] if len(row) > 10 else None
        tags = (
            [t.strip() for t in str(tags_str).split(",") if t.strip()]
            if tags_str
            else None
        )
        # Detect CEFR level from Tags column. If Tags contains A1/A2/B1/B2/C1,
        # use the first such tag as the card's level; otherwise default to B1
        # (this file's original content was all B1).
        level = "B1"
        if tags:
            for t in tags:
                if t in ("A1", "A2", "B1", "B2", "C1"):
                    level = t
                    break
        add_entry(
            bucket,
            row[0],
            categories=tags,
            pos=row[1],
            article=row[2],
            ja=row[3],
            en=row[4],
            prateritum=row[6],
            partizip_ii=row[7],
            plural=row[8],
            notes=row[9],
            level=level,
            source=path.name,
        )


def parse_goethe_online_a1(bucket: dict, path: Path):
    """Goethe DeutschOnline A1 vocabulary (3-line structure: DE / EN / example).

    Strategy:
      1. Filter out chapter/sequence/page headers and copyright lines.
      2. Walk remaining lines in (a, b, c) triplets.
      3. For each window, require:
           - a = plausible German headword (short, not a sentence, not ALL CAPS,
             no preface keywords),
           - b = plausible English translation (no German chars, not a German
             function word, not a German article phrase),
           - and a is not a topic-header sentence.
      4. If the window passes, take the triplet and advance by 3. If not,
         advance by 1 and try again — this skips topic headers that appear
         before each new section.
    """
    SKIP_HEADER = ("DEUTSCH ONLINE A1", "© Goethe-Institut", "GLOSSARY")
    DE_FUNCTION_WORDS = {
        "und", "oder", "aber", "denn", "doch", "auch", "noch", "nur", "schon",
        "sehr", "mehr", "so", "wie", "was", "wo", "wann", "rot", "blau", "gelb",
        "grün", "weiß", "schwarz", "eins", "zwei", "drei", "vier", "fünf", "ja",
        "nein", "bitte", "danke", "mein", "dein", "der", "die", "das", "den",
        "dem", "ein", "eine", "einen", "in", "an", "auf", "bei", "mit", "zu",
        "von", "aus", "nach", "deutsch", "englisch",
    }
    DE_ARTICLE_PREFIX = ("der ", "die ", "das ", "den ", "dem ", "ein ", "eine ")
    DE_PRONOUN_PREFIX = (
        "Ich ", "Du ", "Sie ", "Er ", "Wir ", "Ihr ", "Mein ", "Dein ",
        "ich ", "du ", "wir ", "mein ", "dein ",
    )
    PREFACE_TOKENS = re.compile(
        r"\b(vocabulary|glossary|course|list|chapter|exercise|sequence|presented|"
        r"infinitive|articles|plurals|sentence|example|important|examination|"
        r"core|bold)\b",
        re.IGNORECASE,
    )
    TOPIC_HEADER_PREFIX = (
        "In der ", "In den ", "In dem ", "Bei der ", "Bei den ", "Bei dem ",
        "Zum ", "Zur ", "Über ", "Auf der ", "Auf dem ", "Mit ",
        "Eine ", "Ein ", "Wie ", "Was ", "Wo ", "Woher ", "Wann ", "Welche ",
    )

    def is_skip(line: str) -> bool:
        s = line.strip().lstrip("﻿")
        if not s:
            return True
        if any(p in s for p in SKIP_HEADER):
            return True
        if re.fullmatch(r"\d+", s):
            return True
        if s.startswith(("Kapitel ", "Sequenz ")):
            return True
        return False

    def has_de_chars(s: str) -> bool:
        return bool(re.search(r"[äöüßÄÖÜ]", s))

    def looks_de_term(s: str) -> bool:
        if not s or len(s) > 50:
            return False
        if s.isupper() and len(s) > 3:
            return False
        if s.endswith(":"):
            return False
        if PREFACE_TOKENS.search(s):
            return False
        if re.search(r"[.?!]\s*$", s) and len(s.split()) > 3:
            return False
        return True

    def looks_en(s: str) -> bool:
        if not s or len(s) > 80:
            return False
        if has_de_chars(s):
            return False
        if any(s.startswith(p) for p in DE_ARTICLE_PREFIX + DE_PRONOUN_PREFIX):
            return False
        if s.lower().rstrip(".!?,") in DE_FUNCTION_WORDS:
            return False
        if s.isupper() and len(s) > 3:
            return False
        if s.endswith(":"):
            return False
        if PREFACE_TOKENS.search(s) and len(s.split()) > 3:
            return False
        if re.search(r"[?!]\s*$", s) and len(s.split()) > 3:
            return False
        return True

    def is_topic_header(s: str) -> bool:
        if not s:
            return False
        if has_de_chars(s) or any(s.startswith(p) for p in DE_ARTICLE_PREFIX):
            if re.search(r"[?.!]\s*$", s):
                return True
        if s.startswith(TOPIC_HEADER_PREFIX) and len(s.split()) >= 2:
            return True
        return False

    doc = fitz.open(path)
    all_lines: list[str] = []
    for page in doc:
        for raw in page.get_text().split("\n"):
            s = raw.strip().lstrip("﻿")
            if is_skip(s):
                continue
            all_lines.append(s)
    doc.close()

    i = 0
    while i < len(all_lines) - 2:
        a, b, c = all_lines[i], all_lines[i + 1], all_lines[i + 2]
        if looks_de_term(a) and looks_en(b) and not is_topic_header(a):
            add_entry(
                bucket,
                a,
                en=b,
                example=c if c else None,
                level="A1",
                source=path.name,
            )
            i += 3
        else:
            i += 1


def parse_a1_pdf(bucket: dict, path: Path):
    """NWn_A1_Glossar_Deutsch-Englisch.pdf
    Each vocab is split across 2 lines (DE then EN). Skip headers/footers and
    exercise markers ("12b ÜB", "1", "2a Übung", ...) so that DE/EN pairing
    stays aligned through the page.
    """
    doc = fitz.open(path)
    SKIP_PATTERNS = [
        "Glossar", "Deutsch – Englisch", "© Ernst Klett", "www.klett-sprachen.de",
        "Alle Rechte vorbehalten", "Unterrichtsgebrauch gestattet",
        "Netzwerk neu", "Seite ", "Kapitel ",
    ]
    # Exercise tokens that mark a section header in Klett's layout.
    EXERCISE_TOKENS = {
        "ÜB", "Übung", "Übungen", "Vokabel", "Vokabeln",
        "Schreiben", "Lesen", "Hören", "Sprechen", "Hörverstehen",
        "Grammatik", "Strategie", "Aussprache",
    }
    # Matches a bare ordering label: "1", "2a", "12b" (no other characters).
    NUM_LABEL_RE = re.compile(r"^\s*\d+[a-zA-Z]?\s*$")

    def is_marker(line: str) -> bool:
        if NUM_LABEL_RE.match(line):
            return True
        tokens = line.split()
        if not tokens:
            return True
        # Short header-like lines that contain an exercise token.
        if len(tokens) <= 4 and any(t in EXERCISE_TOKENS for t in tokens):
            return True
        # Page footer fragments: short lines starting with a digit + ALL-CAPS word.
        if len(tokens) == 2 and tokens[0][:1].isdigit() and tokens[1].isupper():
            return True
        return False

    for page in doc:
        txt = page.get_text()
        lines = [L.strip() for L in txt.split("\n") if L.strip()]
        # Filter header/footer text and exercise markers.
        lines = [
            L for L in lines
            if not any(p in L for p in SKIP_PATTERNS) and not is_marker(L)
        ]
        # Pairwise (DE, EN). The page layout puts each vocab entry on two
        # consecutive lines.
        i = 0
        while i < len(lines) - 1:
            de, en = lines[i], lines[i + 1]
            if de and en and not de.isdigit() and len(de) < 80:
                add_entry(bucket, de, en=en, level="A1", source=path.name)
            i += 2
    doc.close()


# ---------- New PDF parsers (B1 / B2 / C1) ----------

_HAS_DE_CHARS = re.compile(r"[äöüßÄÖÜ]")
_DE_ARTICLES_PREFIX = ("der ", "die ", "das ", "den ", "dem ", "ein ", "eine ")
_DE_PRONOUN_PREFIX = ("Ich ", "Du ", "Sie ", "Er ", "Wir ", "Ihr ", "Mein ", "Dein ")


def _strip_paren(s: str) -> str:
    """Remove trailing parenthetical, useful for stripping examples from
    'die Ansicht, -en (Ich bin der Ansicht …)' → 'die Ansicht, -en'."""
    return re.sub(r"\s*\(.*$", "", s).strip()


def _looks_en(line: str) -> bool:
    if not line:
        return False
    if _HAS_DE_CHARS.search(line):
        return False
    if any(line.startswith(p) for p in _DE_ARTICLES_PREFIX + _DE_PRONOUN_PREFIX):
        return False
    return True


def parse_klett_aspekteneu_b1(bucket: dict, path: Path):
    """Klett AspekteNeu B1plus Glossar (German→English).

    Layout: alternating DE block / EN block, each can span 1–2 lines if a
    parenthetical example wraps. EN block detected by ASCII-only content.
    """
    SKIP = [
        "Glossar Deutsch", "Aspekte neu", "Seite ",
        "© Klett", "Klett-Sprachen",
    ]
    SECTION_LABEL = re.compile(r"^\d+[a-z]?$")  # "4a", "1", "12b"
    MODUL_LABEL = re.compile(r"^Modul\s+\d+")

    doc = fitz.open(path)
    lines: list[str] = []
    for page in doc:
        for raw in page.get_text().split("\n"):
            s = raw.strip().lstrip("﻿")
            if not s:
                continue
            if any(p in s for p in SKIP):
                continue
            if SECTION_LABEL.match(s):
                continue
            if MODUL_LABEL.match(s):
                continue
            lines.append(s)
    doc.close()

    # Walk and pair DE block → EN block (each line classified individually).
    i = 0
    while i < len(lines):
        de_buf: list[str] = []
        while i < len(lines) and not _looks_en(lines[i]):
            de_buf.append(lines[i])
            i += 1
        en_buf: list[str] = []
        while i < len(lines) and _looks_en(lines[i]):
            en_buf.append(lines[i])
            i += 1
        if not de_buf or not en_buf:
            continue
        de_full = " ".join(de_buf)
        en_full = " ".join(en_buf)
        # Pull the headword (drop parenthetical example) from each side.
        de_head = _strip_paren(de_full)
        en_head = _strip_paren(en_full)
        if not de_head or len(de_head) > 60:
            continue
        add_entry(
            bucket,
            de_head,
            en=en_head,
            example=de_full if "(" in de_full else None,
            level="B1",
            source=path.name,
        )


def parse_telc_table(bucket: dict, path: Path, level: str):
    """telc Auf jeden Fall / Einfach gut tables.

    Columns repeat: Artikel | Deutsch | Plural | Englisch | Beispielsatz.
    When PyMuPDF flattens, each cell becomes a separate line. We parse by
    looking for an EN line (ASCII, lowercase-ish) and stepping back through
    immediately preceding DE/article lines.

    Robust heuristic: walk lines, accumulate German-looking lines until we
    hit an English line — that's the translation. Then the immediately
    preceding German line is the headword. Any DE article seen in the run
    becomes the article. Any following German line up to the next EN is
    treated as the example sentence.
    """
    SKIP = [
        "telc.net", "Auf jeden Fall", "Einfach gut", "Berufssprachkurse",
        "© telc", "Artikel", "Deutsch", "Plural", "Übersetzung", "Englisch",
        "Beispielsatz", "Lektion ", "Wortschatz zu",
    ]
    ARTICLE_LINE = re.compile(r"^(der|die|das|der/die)$")

    doc = fitz.open(path)
    lines: list[str] = []
    for page in doc:
        for raw in page.get_text().split("\n"):
            s = raw.strip().lstrip("﻿")
            if not s:
                continue
            if any(p in s for p in SKIP):
                continue
            if re.fullmatch(r"\d+", s):
                continue
            lines.append(s)
    doc.close()

    i = 0
    while i < len(lines):
        # Find next EN line — that's a translation candidate.
        while i < len(lines) and not _looks_en(lines[i]):
            i += 1
        if i >= len(lines):
            break
        en_line = lines[i]
        # Look backwards for the matching German headword and article.
        de_term = None
        article = None
        # The line immediately before EN should be the German term.
        j = i - 1
        if j >= 0 and not _looks_en(lines[j]):
            cand = lines[j].strip()
            if not ARTICLE_LINE.match(cand):
                de_term = cand
                # Previous line might be an article.
                if j - 1 >= 0 and ARTICLE_LINE.match(lines[j - 1].strip()):
                    article = lines[j - 1].strip()
        # Look forward for example sentence (German line right after EN).
        example = None
        if i + 1 < len(lines) and not _looks_en(lines[i + 1]):
            cand_ex = lines[i + 1].strip()
            if not ARTICLE_LINE.match(cand_ex) and len(cand_ex) > 5:
                example = cand_ex
        if de_term and 2 <= len(de_term) <= 60:
            add_entry(
                bucket,
                de_term,
                article=article,
                en=en_line,
                example=example,
                level=level,
                source=path.name,
            )
        i += 1


def parse_de_only_list(bucket: dict, path: Path, level: str, skip_patterns: list[str] | None = None):
    """Generic DE-only vocabulary list parser for Klett B2/C1 Kapitelwortschatz,
    Werkstatt C1, and similar. No EN/JA translation captured — these PDFs only
    have German terms with possibly an inline example in parens.

    We split each line at the first '(' to grab just the headword, then validate
    it as a plausible German term and skip the rest.
    """
    SKIP = (skip_patterns or []) + [
        "Kapitelwortschatz", "Aspekte neu", "Werkstatt", "Praxis Spezialverlag",
        "Wortliste", "Seite ", "©",
    ]
    SECTION_LABEL = re.compile(r"^\d+[a-z]?$")
    DE_TERM_RE = re.compile(
        r"^(der|die|das|der/die|die/der)?\s*[A-ZÄÖÜa-zäöüß][\w\-/äöüßÄÖÜ ]*"
    )

    doc = fitz.open(path)
    for page in doc:
        for raw in page.get_text().split("\n"):
            s = raw.strip().lstrip("﻿")
            if not s:
                continue
            if any(p in s for p in SKIP):
                continue
            if SECTION_LABEL.match(s):
                continue
            # Drop parenthetical example to get the headword.
            head = _strip_paren(s)
            head = re.sub(r",\s*[\-A-Za-zäöüß\"]+$", "", head)  # strip plural marker
            head = head.strip()
            if not head or len(head) > 60:
                continue
            if not DE_TERM_RE.match(head):
                continue
            if re.fullmatch(r"[\d\s\W]+", head):
                continue
            # Extract article and term.
            article = None
            term = head
            m = re.match(r"^(der|die|das)\s+(.+)$", head)
            if m:
                article = m.group(1)
                term = m.group(2).strip()
            add_entry(
                bucket,
                term if not article else f"{article} {term}",
                article=article,
                level=level,
                source=path.name,
            )
    doc.close()


def parse_lingster_b2(bucket: dict, path: Path):
    """Lingster Wortschatz bis B2: each entry is 2 lines — DE term + CEFR level
    label (A1 / A2 / B1 / B2). No EN translation.
    """
    LEVEL_LINE = re.compile(r"^(A1|A2|B1|B2|C1)$")
    SKIP = ["Wortschatz", "Seite ", "©", "Lingster"]
    doc = fitz.open(path)
    lines = []
    for page in doc:
        for raw in page.get_text().split("\n"):
            s = raw.strip().lstrip("﻿")
            if not s:
                continue
            if any(p in s for p in SKIP):
                continue
            if re.fullmatch(r"\d+", s):
                continue
            lines.append(s)
    doc.close()

    i = 0
    while i < len(lines) - 1:
        de, lvl = lines[i], lines[i + 1]
        if LEVEL_LINE.match(lvl) and not LEVEL_LINE.match(de):
            head = _strip_paren(de)
            head = re.sub(r",\s*[\-A-Za-zäöüß]+$", "", head).strip()
            if 2 <= len(head) <= 60:
                article = None
                term = head
                m = re.match(r"^(der|die|das)\s+(.+)$", head)
                if m:
                    article = m.group(1)
                    term = m.group(2).strip()
                add_entry(
                    bucket,
                    term if not article else f"{article} {term}",
                    article=article,
                    level=lvl,
                    source=path.name,
                )
            i += 2
        else:
            i += 1


# ---------- Main ----------

def main():
    bucket: dict = {}

    print("Parsing master vocabulary (german_vocab_master.xlsx)...")
    p = SRC_DIR / "german_vocab_master.xlsx"
    if p.exists():
        print(f"  - {p.name}")
        parse_session_new(bucket, p)
    else:
        print(f"  ⚠ master file not found at {p}")

    print("\nAll other sources (B1_xxx.xlsx, German Vocablary.xlsx, all PDFs)")
    print("are intentionally skipped — session_new is the single source of truth.")

    # For any entry without categories from the source, fall back to computed
    # tags. session_new entries already have their Tags column populated, so
    # this branch is effectively a no-op for the master set.
    for entry in bucket.values():
        if not entry["categories"]:
            entry["categories"] = compute_categories(entry)

    # Sort entries by canonical_key for stable output
    entries = sorted(bucket.values(), key=lambda e: e["canonical_key"])

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUT_JSON.open("w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)

    # CSV for human review
    import csv
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "canonical_key", "term_de", "article", "pos",
            "translations_ja", "translations_en",
            "prateritum", "partizip_ii", "plural", "notes",
            "levels", "sources", "categories",
        ])
        for e in entries:
            w.writerow([
                e["canonical_key"],
                e["term_de"],
                e["article"] or "",
                e["pos"] or "",
                " | ".join(e["translations_ja"]),
                " | ".join(e["translations_en"]),
                e["prateritum"] or "",
                e["partizip_ii"] or "",
                e["plural"] or "",
                " | ".join(e["notes"]),
                " | ".join(e["levels"]),
                " | ".join(e["sources"]),
                " | ".join(e["categories"]),
            ])

    # Stats
    print(f"\n{'='*60}")
    print(f"Total unique entries: {len(entries)}")
    by_level = defaultdict(int)
    by_source = defaultdict(int)
    by_category = defaultdict(int)
    has_ja = 0
    has_en = 0
    has_article = 0
    has_pos = 0
    has_any_category = 0
    for e in entries:
        for lv in e["levels"]:
            by_level[lv] += 1
        for s in e["sources"]:
            by_source[s.split("::")[0]] += 1
        for cat in e["categories"]:
            by_category[cat] += 1
        if e["categories"]:
            has_any_category += 1
        if e["translations_ja"]:
            has_ja += 1
        if e["translations_en"]:
            has_en += 1
        if e["article"]:
            has_article += 1
        if e["pos"]:
            has_pos += 1
    print(f"By level: {dict(by_level)}")
    print(f"By source: {dict(by_source)}")
    print(f"By category: {dict(sorted(by_category.items(), key=lambda x: -x[1]))}")
    print(f"With JA translation: {has_ja} ({has_ja * 100 // len(entries)}%)")
    print(f"With EN translation: {has_en} ({has_en * 100 // len(entries)}%)")
    print(f"With at least one category: {has_any_category} ({has_any_category * 100 // len(entries)}%)")
    print(f"With article: {has_article} ({has_article * 100 // len(entries)}%)")
    print(f"With POS: {has_pos} ({has_pos * 100 // len(entries)}%)")
    print(f"\nOutput written to:")
    print(f"  {OUT_JSON}")
    print(f"  {OUT_CSV}")


if __name__ == "__main__":
    main()
