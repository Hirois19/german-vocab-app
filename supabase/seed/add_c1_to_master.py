"""Append C1 vocabulary to german_vocab_session_new.xlsx (the master file).

C1 vocabulary is curated against Goethe Zertifikat C1 / TestDaF / DSH word lists.
Focus: nuanced abstract vocabulary, academic discourse, idiomatic verbs,
formal & literary register, derivational morphology.

Each row carries:
  German term | POS | Article | Japanese | English | Appearance_Count |
  Präteritum  | Partizip II | Plural | Notes | Tags
"""

from __future__ import annotations
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

MASTER = Path(__file__).resolve().parent / "german_vocab_master.xlsx"
SHEET = "Session Vokabeln"
LEVEL = "C1"

C1_ENTRIES: list[tuple[str, str, str, str, str, str, str, str, str]] = [
    # ---------- Nuanced abstract nouns ----------
    ("die Auseinandersetzung", "noun", "die", "対立／取り組み", "dispute, engagement", "", "", "die Auseinandersetzungen", ""),
    ("die Beschäftigung", "noun", "die", "従事／雇用", "occupation, engagement", "", "", "die Beschäftigungen", ""),
    ("der Zusammenstoß", "noun", "der", "衝突", "collision", "", "", "die Zusammenstöße", ""),
    ("die Berücksichtigung", "noun", "die", "考慮", "consideration", "", "", "", ""),
    ("die Verwirklichung", "noun", "die", "実現", "realization", "", "", "die Verwirklichungen", ""),
    ("verwirklichen", "verb", "", "実現する", "to realize", "verwirklichte", "hat verwirklicht", "", ""),
    ("die Verkörperung", "noun", "die", "具現化", "embodiment", "", "", "die Verkörperungen", ""),
    ("verkörpern", "verb", "", "具現化する", "to embody", "verkörperte", "hat verkörpert", "", ""),
    ("die Wahrnehmungsfähigkeit", "noun", "die", "知覚能力", "perceptive ability", "", "", "", ""),
    ("der Spielraum", "noun", "der", "余地", "leeway", "", "", "die Spielräume", ""),
    ("die Reichweite", "noun", "die", "射程／影響範囲", "range, reach", "", "", "die Reichweiten", ""),
    ("die Tragweite", "noun", "die", "重大さ／影響範囲", "magnitude, impact", "", "", "die Tragweiten", ""),
    ("die Tatkraft", "noun", "die", "行動力", "drive", "", "", "", ""),
    ("die Eigenständigkeit", "noun", "die", "自立性", "autonomy", "", "", "", ""),
    ("die Souveränität", "noun", "die", "主権", "sovereignty", "", "", "die Souveränitäten", ""),
    ("die Eigenart", "noun", "die", "独自性", "particularity", "", "", "die Eigenarten", ""),
    ("die Vielfalt", "noun", "die", "多様性", "diversity", "", "", "die Vielfalten", ""),
    ("die Komplexität", "noun", "die", "複雑さ", "complexity", "", "", "die Komplexitäten", ""),
    ("die Schlichtheit", "noun", "die", "簡素さ", "simplicity", "", "", "", ""),
    ("die Geringfügigkeit", "noun", "die", "些細さ", "triviality", "", "", "", ""),
    ("die Beträchtlichkeit", "noun", "die", "重要性", "considerableness", "", "", "", ""),
    ("die Auffassung", "noun", "die", "捉え方", "view, conception", "", "", "die Auffassungen", ""),
    ("die Anschauung", "noun", "die", "見解", "view", "", "", "die Anschauungen", ""),
    ("der Anhaltspunkt", "noun", "der", "手がかり", "clue", "", "", "die Anhaltspunkte", ""),
    ("der Anlass", "noun", "der", "きっかけ", "occasion", "", "", "die Anlässe", ""),
    ("der Auslöser", "noun", "der", "引き金", "trigger", "", "", "die Auslöser", ""),
    ("auslösen", "verb", "", "引き起こす", "to trigger", "löste aus", "hat ausgelöst", "", ""),
    ("der Beweggrund", "noun", "der", "動機", "motive", "", "", "die Beweggründe", ""),
    ("die Motivation", "noun", "die", "動機付け", "motivation", "", "", "die Motivationen", ""),
    ("motivieren", "verb", "", "動機づける", "to motivate", "motivierte", "hat motiviert", "", ""),
    ("der Antrieb", "noun", "der", "原動力", "drive, motivation", "", "", "die Antriebe", ""),
    ("antreiben", "verb", "", "駆り立てる", "to drive", "trieb an", "hat angetrieben", "", "irregular"),
    ("das Streben", "noun", "das", "努力", "striving", "", "", "", ""),
    ("streben", "verb", "", "目指す", "to strive", "strebte", "hat gestrebt", "", "nach"),
    ("das Bemühen", "noun", "das", "努力", "endeavor", "", "", "die Bemühungen", ""),
    ("sich bemühen", "verb", "", "努力する", "to endeavor", "bemühte sich", "hat sich bemüht", "", ""),
    ("die Anstrengung", "noun", "die", "努力", "effort", "", "", "die Anstrengungen", ""),
    ("sich anstrengen", "verb", "", "尽力する", "to exert oneself", "strengte sich an", "hat sich angestrengt", "", ""),
    ("die Hingabe", "noun", "die", "献身", "devotion", "", "", "", ""),
    ("die Einstellung", "noun", "die", "態度／採用", "attitude, hiring", "", "", "die Einstellungen", ""),
    ("die Haltung", "noun", "die", "姿勢／態度", "stance", "", "", "die Haltungen", ""),
    ("die Gesinnung", "noun", "die", "信条", "conviction, mindset", "", "", "die Gesinnungen", ""),
    ("die Weltanschauung", "noun", "die", "世界観", "worldview", "", "", "die Weltanschauungen", ""),
    ("das Selbstverständnis", "noun", "das", "自己認識", "self-perception", "", "", "die Selbstverständnisse", ""),
    ("das Bewusstsein", "noun", "das", "意識", "consciousness", "", "", "", ""),
    ("das Unterbewusstsein", "noun", "das", "潜在意識", "subconscious", "", "", "", ""),
    ("die Erkenntnis", "noun", "die", "認識／知見", "insight", "", "", "die Erkenntnisse", ""),
    ("die Einsicht", "noun", "die", "洞察", "insight", "", "", "die Einsichten", ""),
    ("die Übersicht", "noun", "die", "概観", "overview", "", "", "die Übersichten", ""),
    ("der Überblick", "noun", "der", "全体像", "overview", "", "", "die Überblicke", ""),
    ("die Vorgehensweise", "noun", "die", "進め方", "approach, procedure", "", "", "die Vorgehensweisen", ""),
    ("das Vorgehen", "noun", "das", "進め方", "approach", "", "", "", ""),
    ("die Herangehensweise", "noun", "die", "アプローチ", "approach", "", "", "die Herangehensweisen", ""),
    ("die Maßnahme", "noun", "die", "措置", "measure", "", "", "die Maßnahmen", ""),
    ("ergreifen", "verb", "", "（措置を）取る", "to seize, to take (measures)", "ergriff", "hat ergriffen", "", "irregular"),
    ("der Eingriff", "noun", "der", "介入", "intervention", "", "", "die Eingriffe", ""),
    ("eingreifen", "verb", "", "介入する", "to intervene", "griff ein", "hat eingegriffen", "", "irregular"),
    ("die Intervention", "noun", "die", "介入", "intervention", "", "", "die Interventionen", ""),
    ("intervenieren", "verb", "", "介入する", "to intervene", "intervenierte", "hat interveniert", "", ""),
    ("die Hinsicht", "noun", "die", "観点", "respect, regard", "", "", "die Hinsichten", ""),
    ("der Gesichtspunkt", "noun", "der", "観点", "viewpoint", "", "", "die Gesichtspunkte", ""),
    ("das Ausmaß", "noun", "das", "規模", "extent", "", "", "die Ausmaße", ""),
    ("die Größenordnung", "noun", "die", "規模", "order of magnitude", "", "", "die Größenordnungen", ""),
    ("die Vorgabe", "noun", "die", "指針／目標値", "specification, target", "", "", "die Vorgaben", ""),
    ("die Vorschrift", "noun", "die", "規定", "regulation", "", "", "die Vorschriften", ""),
    ("die Richtlinie", "noun", "die", "ガイドライン", "guideline", "", "", "die Richtlinien", ""),
    ("der Leitfaden", "noun", "der", "ガイド／指針", "guide", "", "", "die Leitfäden", ""),
    ("der Maßstab", "noun", "der", "基準", "standard, scale", "", "", "die Maßstäbe", ""),
    ("das Kriterium", "noun", "das", "基準", "criterion", "", "", "die Kriterien", ""),
    ("der Standard", "noun", "der", "標準", "standard", "", "", "die Standards", ""),
    ("die Norm", "noun", "die", "規範", "norm", "", "", "die Normen", ""),
    ("die Auflage", "noun", "die", "条件／版", "condition, edition", "", "", "die Auflagen", ""),

    # ---------- Academic & analytical verbs ----------
    ("erörtern", "verb", "", "詳しく論じる", "to discuss in detail", "erörterte", "hat erörtert", "", ""),
    ("ausführen", "verb", "", "詳しく述べる", "to elaborate, to execute", "führte aus", "hat ausgeführt", "", ""),
    ("darlegen", "verb", "", "明らかに示す", "to set forth", "legte dar", "hat dargelegt", "", ""),
    ("erläutern", "verb", "", "詳しく説明する", "to elucidate", "erläuterte", "hat erläutert", "", ""),
    ("verdeutlichen", "verb", "", "明確にする", "to clarify", "verdeutlichte", "hat verdeutlicht", "", ""),
    ("präzisieren", "verb", "", "明確化する", "to specify", "präzisierte", "hat präzisiert", "", ""),
    ("differenzieren", "verb", "", "区別する", "to differentiate", "differenzierte", "hat differenziert", "", ""),
    ("nuancieren", "verb", "", "ニュアンスを付ける", "to nuance", "nuancierte", "hat nuanciert", "", ""),
    ("hervorheben", "verb", "", "強調する", "to highlight", "hob hervor", "hat hervorgehoben", "", "irregular"),
    ("betonen", "verb", "", "強調する", "to emphasize", "betonte", "hat betont", "", ""),
    ("unterstreichen", "verb", "", "下線を引く／強調する", "to underline", "unterstrich", "hat unterstrichen", "", "irregular"),
    ("herausstellen", "verb", "", "強調する", "to emphasize", "stellte heraus", "hat herausgestellt", "", ""),
    ("aufzeigen", "verb", "", "明らかにする", "to point out", "zeigte auf", "hat aufgezeigt", "", ""),
    ("offenlegen", "verb", "", "明らかにする", "to disclose", "legte offen", "hat offengelegt", "", ""),
    ("aufdecken", "verb", "", "暴く", "to uncover", "deckte auf", "hat aufgedeckt", "", ""),
    ("enthüllen", "verb", "", "暴露する", "to unveil", "enthüllte", "hat enthüllt", "", ""),
    ("verbergen", "verb", "", "隠す", "to hide", "verbarg", "hat verborgen", "", "irregular"),
    ("verschweigen", "verb", "", "黙秘する", "to keep silent about", "verschwieg", "hat verschwiegen", "", "irregular"),
    ("verleugnen", "verb", "", "否認する", "to deny", "verleugnete", "hat verleugnet", "", ""),
    ("widersprechen", "verb", "", "反論する", "to contradict", "widersprach", "hat widersprochen", "", "irregular"),
    ("einwenden", "verb", "", "異議を唱える", "to object", "wandte ein", "hat eingewandt", "", "mixed"),
    ("der Einwand", "noun", "der", "異議", "objection", "", "", "die Einwände", ""),
    ("entgegnen", "verb", "", "反論する", "to retort", "entgegnete", "hat entgegnet", "", ""),
    ("erwidern", "verb", "", "応答する", "to reply", "erwiderte", "hat erwidert", "", ""),
    ("bestreiten", "verb", "", "争う／否認する", "to dispute", "bestritt", "hat bestritten", "", "irregular"),
    ("anzweifeln", "verb", "", "疑う", "to call into question", "zweifelte an", "hat angezweifelt", "", ""),
    ("hinterfragen", "verb", "", "問いただす", "to question critically", "hinterfragte", "hat hinterfragt", "", ""),
    ("plausibel machen", "verb", "", "もっともらしく示す", "to make plausible", "machte plausibel", "hat plausibel gemacht", "", ""),
    ("rechtfertigen", "verb", "", "正当化する", "to justify", "rechtfertigte", "hat gerechtfertigt", "", ""),
    ("die Rechtfertigung", "noun", "die", "正当化", "justification", "", "", "die Rechtfertigungen", ""),
    ("legitimieren", "verb", "", "正当化する", "to legitimize", "legitimierte", "hat legitimiert", "", ""),
    ("die Legitimation", "noun", "die", "正当性", "legitimation", "", "", "die Legitimationen", ""),
    ("vorgeben", "verb", "", "見せかける／指示する", "to pretend, to specify", "gab vor", "hat vorgegeben", "", "irregular"),
    ("vortäuschen", "verb", "", "ふりをする", "to feign", "täuschte vor", "hat vorgetäuscht", "", ""),
    ("die Täuschung", "noun", "die", "錯覚／欺き", "deception", "", "", "die Täuschungen", ""),
    ("täuschen", "verb", "", "騙す", "to deceive", "täuschte", "hat getäuscht", "", ""),
    ("die Illusion", "noun", "die", "幻想", "illusion", "", "", "die Illusionen", ""),
    ("desillusionieren", "verb", "", "幻滅させる", "to disillusion", "desillusionierte", "hat desillusioniert", "", ""),
    ("verdeutlichen", "verb", "", "明確にする", "to make clear", "verdeutlichte", "hat verdeutlicht", "", ""),
    ("konkretisieren", "verb", "", "具体化する", "to concretize", "konkretisierte", "hat konkretisiert", "", ""),
    ("verallgemeinern", "verb", "", "一般化する", "to generalize", "verallgemeinerte", "hat verallgemeinert", "", ""),
    ("die Verallgemeinerung", "noun", "die", "一般化", "generalization", "", "", "die Verallgemeinerungen", ""),
    ("abstrahieren", "verb", "", "抽象化する", "to abstract", "abstrahierte", "hat abstrahiert", "", ""),
    ("die Abstraktion", "noun", "die", "抽象化", "abstraction", "", "", "die Abstraktionen", ""),
    ("präzisieren", "verb", "", "明確化する", "to specify", "präzisierte", "hat präzisiert", "", ""),
    ("die Präzision", "noun", "die", "正確性", "precision", "", "", "", ""),
    ("variieren", "verb", "", "変化させる", "to vary", "variierte", "hat variiert", "", ""),
    ("die Variation", "noun", "die", "変動", "variation", "", "", "die Variationen", ""),
    ("die Variante", "noun", "die", "変種", "variant", "", "", "die Varianten", ""),
    ("modifizieren", "verb", "", "修正する", "to modify", "modifizierte", "hat modifiziert", "", ""),
    ("die Modifikation", "noun", "die", "修正", "modification", "", "", "die Modifikationen", ""),
    ("anpassen", "verb", "", "適応させる", "to adapt", "passte an", "hat angepasst", "", ""),
    ("die Anpassung", "noun", "die", "適応", "adaptation", "", "", "die Anpassungen", ""),
    ("revidieren", "verb", "", "改訂する", "to revise", "revidierte", "hat revidiert", "", ""),
    ("die Revision", "noun", "die", "改訂", "revision", "", "", "die Revisionen", ""),
    ("überarbeiten", "verb", "", "改訂する", "to revise", "überarbeitete", "hat überarbeitet", "", ""),
    ("die Überarbeitung", "noun", "die", "改訂", "revision", "", "", "die Überarbeitungen", ""),

    # ---------- Society advanced ----------
    ("die Demografie", "noun", "die", "人口統計学", "demography", "", "", "", ""),
    ("demografisch", "adj", "", "人口統計学的な", "demographic", "", "", "", ""),
    ("der demografische Wandel", "noun", "der", "人口動態の変化", "demographic change", "", "", "", ""),
    ("die Alterung", "noun", "die", "高齢化", "aging", "", "", "", ""),
    ("die Überalterung", "noun", "die", "高齢化", "overaging", "", "", "", ""),
    ("die Geburtenrate", "noun", "die", "出生率", "birth rate", "", "", "die Geburtenraten", ""),
    ("die Sterblichkeitsrate", "noun", "die", "死亡率", "mortality rate", "", "", "die Sterblichkeitsraten", ""),
    ("die Migration", "noun", "die", "移住", "migration", "", "", "die Migrationen", ""),
    ("migrieren", "verb", "", "移住する", "to migrate", "migrierte", "ist migriert", "", ""),
    ("der Migrant", "noun", "der", "移民", "migrant", "", "", "die Migranten", "n-decl"),
    ("die Auswanderung", "noun", "die", "移住（出国）", "emigration", "", "", "die Auswanderungen", ""),
    ("die Einwanderung", "noun", "die", "移住（入国）", "immigration", "", "", "die Einwanderungen", ""),
    ("die Flucht", "noun", "die", "逃走", "flight, escape", "", "", "die Fluchten", ""),
    ("fliehen", "verb", "", "逃げる", "to flee", "floh", "ist geflohen", "", "irregular"),
    ("der Asylant", "noun", "der", "亡命者", "asylum seeker", "", "", "die Asylanten", "n-decl"),
    ("das Asyl", "noun", "das", "庇護", "asylum", "", "", "die Asyle", ""),
    ("die Toleranz", "noun", "die", "寛容", "tolerance", "", "", "die Toleranzen", ""),
    ("tolerieren", "verb", "", "容認する", "to tolerate", "tolerierte", "hat toleriert", "", ""),
    ("die Intoleranz", "noun", "die", "不寛容", "intolerance", "", "", "die Intoleranzen", ""),
    ("die Vielfalt", "noun", "die", "多様性", "diversity", "", "", "die Vielfalten", ""),
    ("die Diversität", "noun", "die", "多様性", "diversity", "", "", "die Diversitäten", ""),
    ("der Pluralismus", "noun", "der", "多元主義", "pluralism", "", "", "", ""),
    ("die Ungleichheit", "noun", "die", "不平等", "inequality", "", "", "die Ungleichheiten", ""),
    ("die Ungerechtigkeit", "noun", "die", "不公正", "injustice", "", "", "die Ungerechtigkeiten", ""),
    ("die Gerechtigkeit", "noun", "die", "正義", "justice", "", "", "", ""),
    ("die Solidarität", "noun", "die", "連帯", "solidarity", "", "", "", ""),
    ("solidarisch", "adj", "", "連帯した", "in solidarity", "", "", "", ""),
    ("der Wohlstand", "noun", "der", "繁栄", "prosperity", "", "", "", ""),
    ("die Armut", "noun", "die", "貧困", "poverty", "", "", "", ""),
    ("der Reichtum", "noun", "der", "富", "wealth", "", "", "", ""),
    ("die Mittelschicht", "noun", "die", "中流階級", "middle class", "", "", "die Mittelschichten", ""),
    ("die Oberschicht", "noun", "die", "上流階級", "upper class", "", "", "die Oberschichten", ""),
    ("die Unterschicht", "noun", "die", "下流階級", "lower class", "", "", "die Unterschichten", ""),
    ("die Schicht", "noun", "die", "階層", "layer, class", "", "", "die Schichten", ""),
    ("die Generation", "noun", "die", "世代", "generation", "", "", "die Generationen", ""),
    ("der Generationenkonflikt", "noun", "der", "世代間対立", "generational conflict", "", "", "die Generationenkonflikte", ""),
    ("die Emanzipation", "noun", "die", "解放", "emancipation", "", "", "die Emanzipationen", ""),
    ("die Gleichstellung", "noun", "die", "対等な地位", "equality", "", "", "die Gleichstellungen", ""),
    ("die Geschlechtergerechtigkeit", "noun", "die", "ジェンダー公正", "gender justice", "", "", "", ""),
    ("die Feminismus", "noun", "der", "フェミニズム", "feminism", "", "", "", "der Feminismus"),

    # ---------- Politics advanced ----------
    ("die Außenpolitik", "noun", "die", "外交政策", "foreign policy", "", "", "", ""),
    ("die Innenpolitik", "noun", "die", "国内政策", "domestic policy", "", "", "", ""),
    ("die Sozialpolitik", "noun", "die", "社会政策", "social policy", "", "", "", ""),
    ("die Wirtschaftspolitik", "noun", "die", "経済政策", "economic policy", "", "", "", ""),
    ("die Bildungspolitik", "noun", "die", "教育政策", "education policy", "", "", "", ""),
    ("die Umweltpolitik", "noun", "die", "環境政策", "environmental policy", "", "", "", ""),
    ("das Bündnis", "noun", "das", "同盟", "alliance", "", "", "die Bündnisse", ""),
    ("die Koalition", "noun", "die", "連立", "coalition", "", "", "die Koalitionen", ""),
    ("die Opposition", "noun", "die", "野党", "opposition", "", "", "die Oppositionen", ""),
    ("die Mehrheit", "noun", "die", "多数", "majority", "", "", "die Mehrheiten", ""),
    ("die Minderheit", "noun", "die", "少数", "minority", "", "", "die Minderheiten", ""),
    ("die Stimme", "noun", "die", "投票／声", "vote, voice", "", "", "die Stimmen", ""),
    ("die Stimmenmehrheit", "noun", "die", "多数票", "majority of votes", "", "", "", ""),
    ("die Volksabstimmung", "noun", "die", "国民投票", "referendum", "", "", "die Volksabstimmungen", ""),
    ("das Referendum", "noun", "das", "国民投票", "referendum", "", "", "die Referenden", ""),
    ("die Verfassungsänderung", "noun", "die", "憲法改正", "constitutional amendment", "", "", "die Verfassungsänderungen", ""),
    ("die Gesetzgebung", "noun", "die", "立法", "legislation", "", "", "", ""),
    ("der Gesetzentwurf", "noun", "der", "法案", "bill", "", "", "die Gesetzentwürfe", ""),
    ("verabschieden", "verb", "", "可決する／別れる", "to pass (a law), to bid farewell", "verabschiedete", "hat verabschiedet", "", ""),
    ("inkrafttreten", "verb", "", "施行される", "to come into force", "trat in Kraft", "ist in Kraft getreten", "", "irregular"),
    ("der Vertreter", "noun", "der", "代表者", "representative", "", "", "die Vertreter", ""),
    ("die Vertreterin", "noun", "die", "代表者（女）", "representative (f)", "", "", "die Vertreterinnen", ""),
    ("das Mandat", "noun", "das", "委任", "mandate", "", "", "die Mandate", ""),
    ("die Amtszeit", "noun", "die", "任期", "term of office", "", "", "die Amtszeiten", ""),
    ("die Wahlkampagne", "noun", "die", "選挙運動", "election campaign", "", "", "die Wahlkampagnen", ""),
    ("die Lobby", "noun", "die", "ロビー", "lobby", "", "", "die Lobbys", ""),
    ("der Skandal", "noun", "der", "スキャンダル", "scandal", "", "", "die Skandale", ""),
    ("die Korruption", "noun", "die", "汚職", "corruption", "", "", "die Korruptionen", ""),
    ("korrupt", "adj", "", "汚職した", "corrupt", "", "", "", ""),
    ("die Bestechung", "noun", "die", "賄賂", "bribery", "", "", "die Bestechungen", ""),
    ("bestechen", "verb", "", "賄賂を渡す", "to bribe", "bestach", "hat bestochen", "", "irregular"),

    # ---------- Idiomatic & nuanced expressions ----------
    ("sich auseinandersetzen", "verb", "", "取り組む／対立する", "to engage with, to argue with", "setzte sich auseinander", "hat sich auseinandergesetzt", "", ""),
    ("sich befassen", "verb", "", "取り組む", "to deal with", "befasste sich", "hat sich befasst", "", "+ mit"),
    ("sich beschäftigen", "verb", "", "従事する", "to occupy oneself with", "beschäftigte sich", "hat sich beschäftigt", "", "+ mit"),
    ("sich abfinden", "verb", "", "受け入れる", "to come to terms with", "fand sich ab", "hat sich abgefunden", "", "irregular + mit"),
    ("sich verständigen", "verb", "", "意思疎通する", "to communicate", "verständigte sich", "hat sich verständigt", "", "+ mit"),
    ("sich anpassen", "verb", "", "適応する", "to adapt", "passte sich an", "hat sich angepasst", "", "+ an"),
    ("sich gewöhnen", "verb", "", "慣れる", "to get used to", "gewöhnte sich", "hat sich gewöhnt", "", "+ an"),
    ("sich orientieren", "verb", "", "方向づける", "to orient oneself", "orientierte sich", "hat sich orientiert", "", "+ an"),
    ("sich richten", "verb", "", "向かう／合わせる", "to be directed at", "richtete sich", "hat sich gerichtet", "", "+ an"),
    ("sich beziehen", "verb", "", "言及する", "to refer to", "bezog sich", "hat sich bezogen", "", "+ auf"),
    ("sich erweisen", "verb", "", "～と判明する", "to turn out to be", "erwies sich", "hat sich erwiesen", "", "+ als"),
    ("sich herausstellen", "verb", "", "明らかになる", "to turn out", "stellte sich heraus", "hat sich herausgestellt", "", ""),
    ("sich ergeben", "verb", "", "結果として生じる", "to result, to surrender", "ergab sich", "hat sich ergeben", "", "irregular"),
    ("sich abzeichnen", "verb", "", "徐々に現れる", "to emerge", "zeichnete sich ab", "hat sich abgezeichnet", "", ""),
    ("sich ankündigen", "verb", "", "予告される", "to be announced", "kündigte sich an", "hat sich angekündigt", "", ""),
    ("zur Sprache kommen", "phrase", "", "話題に上る", "to come up (as topic)", "kam zur Sprache", "ist zur Sprache gekommen", "", ""),
    ("zur Anwendung kommen", "phrase", "", "適用される", "to be applied", "", "", "", ""),
    ("in Erwägung ziehen", "phrase", "", "考慮に入れる", "to take into consideration", "", "", "", ""),
    ("in Frage stellen", "phrase", "", "疑問視する", "to call into question", "", "", "", ""),
    ("in Kauf nehmen", "phrase", "", "甘受する", "to accept (downside)", "", "", "", ""),
    ("in Erscheinung treten", "phrase", "", "現れる", "to appear", "", "", "", ""),
    ("zur Folge haben", "phrase", "", "結果を持つ", "to result in", "", "", "", ""),
    ("zur Geltung kommen", "phrase", "", "活きる／際立つ", "to come into its own", "", "", "", ""),
    ("Bezug nehmen", "phrase", "", "言及する", "to refer to", "", "", "", "+ auf"),
    ("Stellung nehmen", "phrase", "", "立場を表明する", "to take a position", "", "", "", ""),
    ("zu dem Schluss kommen", "phrase", "", "結論に達する", "to come to the conclusion", "", "", "", ""),

    # ---------- Higher register adjectives ----------
    ("beträchtlich", "adj", "", "相当な", "considerable", "", "", "", ""),
    ("erheblich", "adj", "", "著しい", "significant", "", "", "", ""),
    ("maßgeblich", "adj", "", "決定的な", "authoritative", "", "", "", ""),
    ("ausschlaggebend", "adj", "", "決定的な", "decisive", "", "", "", ""),
    ("zwingend", "adj", "", "強制的な", "compelling", "", "", "", ""),
    ("unumgänglich", "adj", "", "避けられない", "unavoidable", "", "", "", ""),
    ("unentbehrlich", "adj", "", "不可欠な", "indispensable", "", "", "", ""),
    ("unverzichtbar", "adj", "", "なくてはならない", "indispensable", "", "", "", ""),
    ("unabdingbar", "adj", "", "必要不可欠な", "indispensable", "", "", "", ""),
    ("zwangsläufig", "adj", "", "必然的な", "inevitable", "", "", "", ""),
    ("unvermeidlich", "adj", "", "避けられない", "inevitable", "", "", "", ""),
    ("gravierend", "adj", "", "深刻な", "serious", "", "", "", ""),
    ("tiefgreifend", "adj", "", "根本的な", "profound", "", "", "", ""),
    ("weitreichend", "adj", "", "広範な", "far-reaching", "", "", "", ""),
    ("umfassend", "adj", "", "包括的な", "comprehensive", "", "", "", ""),
    ("ganzheitlich", "adj", "", "全体論的な", "holistic", "", "", "", ""),
    ("nachhaltig", "adj", "", "持続可能な／持続的な", "sustainable, lasting", "", "", "", ""),
    ("dauerhaft", "adj", "", "永続的な", "lasting", "", "", "", ""),
    ("vorübergehend", "adj", "", "一時的な", "temporary", "", "", "", ""),
    ("kurzfristig", "adj", "", "短期的な", "short-term", "", "", "", ""),
    ("langfristig", "adj", "", "長期的な", "long-term", "", "", "", ""),
    ("mittelfristig", "adj", "", "中期的な", "medium-term", "", "", "", ""),
    ("absehbar", "adj", "", "予見可能な", "foreseeable", "", "", "", ""),
    ("unvorhersehbar", "adj", "", "予見不可能な", "unforeseeable", "", "", "", ""),
    ("schwerwiegend", "adj", "", "重大な", "grave", "", "", "", ""),
    ("ausgeprägt", "adj", "", "顕著な", "pronounced", "", "", "", ""),
    ("herausragend", "adj", "", "傑出した", "outstanding", "", "", "", ""),
    ("hervorragend", "adj", "", "卓越した", "excellent", "", "", "", ""),
    ("hervorstechend", "adj", "", "目立つ", "salient", "", "", "", ""),
    ("auffällig", "adj", "", "目立つ", "conspicuous", "", "", "", ""),
    ("unauffällig", "adj", "", "目立たない", "inconspicuous", "", "", "", ""),
    ("bemerkenswert", "adj", "", "注目すべき", "remarkable", "", "", "", ""),
    ("beachtenswert", "adj", "", "注目に値する", "noteworthy", "", "", "", ""),
    ("erstaunlich", "adj", "", "驚くべき", "astonishing", "", "", "", ""),
    ("verblüffend", "adj", "", "驚異的な", "astonishing", "", "", "", ""),
    ("verwunderlich", "adj", "", "驚くべき", "surprising", "", "", "", ""),
    ("unbestreitbar", "adj", "", "否定できない", "indisputable", "", "", "", ""),
    ("unstrittig", "adj", "", "議論の余地のない", "uncontroversial", "", "", "", ""),
    ("strittig", "adj", "", "議論のある", "controversial", "", "", "", ""),
    ("kontrovers", "adj", "", "論争的な", "controversial", "", "", "", ""),
    ("brisant", "adj", "", "微妙な／緊迫した", "explosive, sensitive", "", "", "", ""),
    ("heikel", "adj", "", "デリケートな", "delicate", "", "", "", ""),
    ("delikat", "adj", "", "デリケートな", "delicate", "", "", "", ""),
    ("subtil", "adj", "", "繊細な", "subtle", "", "", "", ""),
    ("differenziert", "adj", "", "差別化された／詳細な", "differentiated, nuanced", "", "", "", ""),
    ("nuanciert", "adj", "", "ニュアンスのある", "nuanced", "", "", "", ""),
    ("plausibel", "adj", "", "もっともらしい", "plausible", "", "", "", ""),
    ("nachvollziehbar", "adj", "", "理解可能な", "comprehensible", "", "", "", ""),
    ("schlüssig", "adj", "", "首尾一貫した", "conclusive", "", "", "", ""),
    ("stichhaltig", "adj", "", "確かな", "valid", "", "", "", ""),
    ("triftig", "adj", "", "もっともな", "valid", "", "", "", ""),
    ("zutreffend", "adj", "", "当てはまる", "applicable, accurate", "", "", "", ""),
    ("unzutreffend", "adj", "", "当てはまらない", "inaccurate", "", "", "", ""),
    ("paradox", "adj", "", "逆説的な", "paradoxical", "", "", "", ""),
    ("widersprüchlich", "adj", "", "矛盾した", "contradictory", "", "", "", ""),
    ("konsistent", "adj", "", "一貫した", "consistent", "", "", "", ""),
    ("inkonsistent", "adj", "", "一貫しない", "inconsistent", "", "", "", ""),
    ("kohärent", "adj", "", "首尾一貫した", "coherent", "", "", "", ""),
    ("inkohärent", "adj", "", "首尾一貫しない", "incoherent", "", "", "", ""),
    ("gewissermaßen", "adv", "", "いわば", "to some extent", "", "", "", ""),
    ("gleichsam", "adv", "", "いわば", "as if", "", "", "", ""),
    ("sozusagen", "adv", "", "いわば", "so to speak", "", "", "", ""),
    ("vermeintlich", "adj/adv", "", "そう思われる", "supposed", "", "", "", ""),
    ("angeblich", "adj/adv", "", "～と言われる", "allegedly", "", "", "", ""),
    ("vorgeblich", "adj/adv", "", "見せかけの", "ostensibly", "", "", "", ""),
    ("scheinbar", "adj/adv", "", "見かけ上の", "apparent", "", "", "", ""),
    ("anscheinend", "adv", "", "どうやら", "apparently", "", "", "", ""),
    ("offenkundig", "adj", "", "明白な", "evident", "", "", "", ""),
    ("augenscheinlich", "adj/adv", "", "明らかな", "evidently", "", "", "", ""),

    # ---------- C1 connectors & discourse markers ----------
    ("ungeachtet", "prep", "", "～にもかかわらず", "regardless of", "", "", "", "+Gen"),
    ("dessen ungeachtet", "phrase", "", "それにもかかわらず", "regardless", "", "", "", ""),
    ("nichtsdestotrotz", "adv", "", "それにもかかわらず", "nonetheless", "", "", "", ""),
    ("nichtsdestoweniger", "adv", "", "それにもかかわらず", "nevertheless", "", "", "", ""),
    ("gleichwohl", "adv", "", "それでもなお", "yet, however", "", "", "", ""),
    ("indes", "conj", "", "とはいえ", "however", "", "", "", ""),
    ("insofern", "conj", "", "～する限りでは", "insofar as", "", "", "", ""),
    ("insoweit", "conj", "", "～する限りでは", "to that extent", "", "", "", ""),
    ("insbesondere", "adv", "", "とりわけ", "in particular", "", "", "", ""),
    ("vornehmlich", "adv", "", "主に", "primarily", "", "", "", ""),
    ("überwiegend", "adv", "", "主として", "predominantly", "", "", "", ""),
    ("vorwiegend", "adv", "", "主に", "mainly", "", "", "", ""),
    ("hauptsächlich", "adv", "", "主に", "mainly", "", "", "", ""),
    ("primär", "adv", "", "第一に", "primarily", "", "", "", ""),
    ("sekundär", "adv", "", "二次的に", "secondarily", "", "", "", ""),
    ("hinzu kommt, dass", "phrase", "", "さらに～", "in addition", "", "", "", ""),
    ("um so mehr, als", "phrase", "", "なおさら～", "all the more, as", "", "", "", ""),
    ("umso", "adv", "", "ますます", "all the more", "", "", "", ""),
    ("im Folgenden", "phrase", "", "以下において", "in the following", "", "", "", ""),
    ("im Wesentlichen", "phrase", "", "本質的に", "essentially", "", "", "", ""),
    ("im Großen und Ganzen", "phrase", "", "おおまかに言って", "by and large", "", "", "", ""),
    ("alles in allem", "phrase", "", "全体として", "all in all", "", "", "", ""),
    ("letzten Endes", "phrase", "", "結局", "in the end", "", "", "", ""),
    ("schlussendlich", "adv", "", "最終的に", "ultimately", "", "", "", ""),
    ("im Endeffekt", "phrase", "", "結局のところ", "in the end", "", "", "", ""),
    ("letztlich", "adv", "", "最終的に", "ultimately", "", "", "", ""),
    ("unter dem Strich", "phrase", "", "結論として", "at the end of the day", "", "", "", ""),
    ("im Übrigen", "phrase", "", "ところで", "by the way, incidentally", "", "", "", ""),
    ("nebenbei bemerkt", "phrase", "", "ちなみに", "by the way", "", "", "", ""),
    ("am Rande bemerkt", "phrase", "", "余談だが", "as an aside", "", "", "", ""),
    ("anbei", "adv", "", "添付", "enclosed", "", "", "", ""),
    ("hierbei", "adv", "", "この際", "in this", "", "", "", ""),
    ("dabei", "adv", "", "その際", "thereby, at the same time", "", "", "", ""),
    ("hiermit", "adv", "", "ここに", "hereby", "", "", "", ""),
    ("darin", "adv", "", "それに", "in it", "", "", "", ""),
    ("daraus", "adv", "", "それから", "from it", "", "", "", ""),
    ("dazu", "adv", "", "それに", "in addition", "", "", "", ""),
    ("davon", "adv", "", "そこから", "of it", "", "", "", ""),
    ("davor", "adv", "", "その前に", "before that", "", "", "", ""),
    ("danach", "adv", "", "その後", "after that", "", "", "", ""),
    ("dadurch", "adv", "", "それによって", "thereby", "", "", "", ""),
    ("dafür", "adv", "", "そのために", "for that", "", "", "", ""),
    ("dagegen", "adv", "", "それに対して", "against it", "", "", "", ""),
    ("darauf", "adv", "", "その上に", "on it", "", "", "", ""),
    ("darüber", "adv", "", "その上", "above it", "", "", "", ""),
    ("darunter", "adv", "", "その下に", "below it", "", "", "", ""),
    ("worüber", "adv", "", "何について", "about what", "", "", "", ""),
    ("wovon", "adv", "", "何から", "of what", "", "", "", ""),
    ("worauf", "adv", "", "何の上に", "on what", "", "", "", ""),
    ("worin", "adv", "", "何の中に", "in what", "", "", "", ""),

    # ---------- Formal & literary nouns ----------
    ("die Etablierung", "noun", "die", "確立", "establishment", "", "", "die Etablierungen", ""),
    ("etablieren", "verb", "", "確立する", "to establish", "etablierte", "hat etabliert", "", ""),
    ("die Implementierung", "noun", "die", "実装", "implementation", "", "", "die Implementierungen", ""),
    ("implementieren", "verb", "", "実装する", "to implement", "implementierte", "hat implementiert", "", ""),
    ("die Umsetzung", "noun", "die", "実施", "implementation", "", "", "die Umsetzungen", ""),
    ("umsetzen", "verb", "", "実行する", "to implement", "setzte um", "hat umgesetzt", "", ""),
    ("die Durchführung", "noun", "die", "実施", "execution", "", "", "die Durchführungen", ""),
    ("durchführen", "verb", "", "実施する", "to carry out", "führte durch", "hat durchgeführt", "", ""),
    ("die Verwirklichung", "noun", "die", "実現", "realization", "", "", "die Verwirklichungen", ""),
    ("die Realisierung", "noun", "die", "実現", "realization", "", "", "die Realisierungen", ""),
    ("realisieren", "verb", "", "実現する／気づく", "to realize", "realisierte", "hat realisiert", "", ""),
    ("die Konsolidierung", "noun", "die", "強化／統合", "consolidation", "", "", "die Konsolidierungen", ""),
    ("konsolidieren", "verb", "", "強化する", "to consolidate", "konsolidierte", "hat konsolidiert", "", ""),
    ("die Stagnation", "noun", "die", "停滞", "stagnation", "", "", "die Stagnationen", ""),
    ("stagnieren", "verb", "", "停滞する", "to stagnate", "stagnierte", "hat stagniert", "", ""),
    ("die Stabilisierung", "noun", "die", "安定化", "stabilization", "", "", "die Stabilisierungen", ""),
    ("stabilisieren", "verb", "", "安定化する", "to stabilize", "stabilisierte", "hat stabilisiert", "", ""),
    ("die Destabilisierung", "noun", "die", "不安定化", "destabilization", "", "", "die Destabilisierungen", ""),
    ("die Eskalation", "noun", "die", "エスカレーション", "escalation", "", "", "die Eskalationen", ""),
    ("eskalieren", "verb", "", "エスカレートする", "to escalate", "eskalierte", "ist eskaliert", "", ""),
    ("die Deeskalation", "noun", "die", "デエスカレーション", "de-escalation", "", "", "die Deeskalationen", ""),
    ("die Beilegung", "noun", "die", "解決", "settlement", "", "", "die Beilegungen", ""),
    ("beilegen", "verb", "", "解決する", "to settle", "legte bei", "hat beigelegt", "", ""),
    ("die Schlichtung", "noun", "die", "調停", "mediation", "", "", "die Schlichtungen", ""),
    ("schlichten", "verb", "", "調停する", "to mediate", "schlichtete", "hat geschlichtet", "", ""),
    ("die Vermittlung", "noun", "die", "仲介", "mediation", "", "", "die Vermittlungen", ""),
    ("vermitteln", "verb", "", "仲介する／伝える", "to mediate, to convey", "vermittelte", "hat vermittelt", "", ""),
    ("der Vermittler", "noun", "der", "仲介者", "mediator", "", "", "die Vermittler", ""),
]

THEMES = {
    "Familie": ["family", "Familie", "Vater", "Mutter", "Kind"],
    "Arbeit": ["work", "job", "Arbeit", "Beruf", "Karriere", "Unternehmer", "Vertreter"],
    "Wirtschaft": [
        "economy", "market", "trade", "investment", "growth",
        "経済", "市場", "投資",
        "Wirtschaft", "Markt", "Stagnation", "Stabilisierung", "Inflation",
        "Konsolidierung",
    ],
    "Politik": [
        "politics", "government", "election", "law", "policy", "democracy",
        "constitution",
        "政治", "政府", "選挙", "法律", "政策", "民主主義", "憲法",
        "Politik", "Außenpolitik", "Innenpolitik", "Sozialpolitik",
        "Wirtschaftspolitik", "Bildungspolitik", "Umweltpolitik", "Regierung",
        "Wahl", "Partei", "Gesetz", "Koalition", "Opposition", "Stimme",
        "Mandat", "Amtszeit", "Wahlkampagne", "Lobby", "Korruption",
        "Bestechung", "Vertreter", "Verfassung", "Verfassungsänderung",
        "Gesetzgebung", "Gesetzentwurf", "Volksabstimmung", "Referendum",
        "Skandal", "Bündnis",
    ],
    "Gesellschaft": [
        "society", "social", "migration", "tolerance", "equality", "minority",
        "diversity", "generation", "emancipation",
        "社会", "移住", "寛容", "平等", "少数派", "多様性", "世代", "貧困",
        "Gesellschaft", "Demografie", "demografisch", "Wandel", "Alterung",
        "Überalterung", "Geburtenrate", "Sterblichkeitsrate", "Migration",
        "Migrant", "Auswanderung", "Einwanderung", "Flucht", "Asyl",
        "Asylant", "Toleranz", "tolerieren", "Intoleranz", "Vielfalt",
        "Diversität", "Pluralismus", "Ungleichheit", "Ungerechtigkeit",
        "Gerechtigkeit", "Solidarität", "solidarisch", "Wohlstand", "Armut",
        "Reichtum", "Mittelschicht", "Oberschicht", "Unterschicht", "Schicht",
        "Generation", "Generationenkonflikt", "Emanzipation", "Gleichstellung",
        "Feminismus",
    ],
    "Wissenschaft": [
        "science", "research", "analysis", "theory", "academic",
        "科学", "研究", "分析", "理論", "学術",
        "Wissenschaft", "Forschung", "analysieren", "Theorie", "Erkenntnis",
        "Einsicht", "Abstraktion", "abstrahieren", "konkretisieren",
        "Modifikation", "Variation", "Revision",
    ],
    "Medien": [
        "media", "press", "publication", "media", "newspaper",
        "メディア", "報道", "出版",
        "Medien", "Presse", "Veröffentlichung", "Zitat", "Pressemitteilung",
        "Stellungnahme", "Skandal",
    ],
    "Gefühle": [
        "feeling", "emotion", "joy", "sad", "stress",
        "感情", "喜び", "悲しみ",
        "Gefühl", "Emotion", "Hoffnung", "Sehnsucht", "Sympathie", "Empathie",
        "Charakter", "Persönlichkeit",
    ],
    "Umwelt": [
        "environment", "climate", "sustainable", "energy",
        "環境", "気候", "持続可能",
        "Umwelt", "Klima", "nachhaltig", "Nachhaltigkeit",
    ],
    "Bildung": [
        "education", "university", "study", "exam",
        "教育", "大学", "学術",
        "Bildung", "Universität", "Studium", "akademisch",
    ],
}

_ASCII_WORD = re.compile(r"^[a-zA-Z]+$")


def has_keyword(haystack: str, keyword: str) -> bool:
    if _ASCII_WORD.fullmatch(keyword):
        return bool(re.search(rf"\b{re.escape(keyword)}\b", haystack, re.IGNORECASE))
    return keyword in haystack


def compute_tags(german: str, pos: str, ja: str, en: str, level: str) -> list[str]:
    haystack = " ".join(filter(None, [german, ja, en]))
    tags: list[str] = []
    for theme, kws in THEMES.items():
        for kw in kws:
            if has_keyword(haystack, kw):
                tags.append(theme)
                break
    if pos:
        tags.append(pos.strip().lower())
    if level:
        tags.append(level)
    seen: set[str] = set()
    out = []
    for t in tags:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


def main():
    wb = load_workbook(MASTER)
    ws = wb[SHEET]
    header = [c.value for c in ws[1]]
    if "Tags" not in header:
        raise RuntimeError("Tags column missing")

    existing_terms: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        g = row[0]
        if g and str(g).strip():
            existing_terms.add(str(g).strip().lower())

    start_row = ws.max_row + 1
    added = 0
    skipped = 0
    for g, pos, art, ja, en, prat, pii, plural, notes in C1_ENTRIES:
        if g.strip().lower() in existing_terms:
            skipped += 1
            continue
        tags = compute_tags(g, pos, ja, en, LEVEL)
        row_data = [g, pos, art, ja, en, 1, prat, pii, plural, notes, ", ".join(tags)]
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=start_row + added, column=col_idx, value=value)
        added += 1
        existing_terms.add(g.strip().lower())

    wb.save(MASTER)
    print(f"Added {added} C1 entries. Skipped {skipped} (already present). Total rows now: {ws.max_row}.")


if __name__ == "__main__":
    main()
